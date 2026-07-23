import streamlit as st
import pandas as pd
import plotly.express as px
import psycopg2
import psycopg2.extras
import calendar
from datetime import datetime, timedelta
import json
import io
import smtplib
from email.message import EmailMessage
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import time

# Performance optimization: Cache expensive computations
@st.cache_data(ttl=3600, show_spinner=False)
def generate_formatted_excel(df, subtotal_indices=None):
    if subtotal_indices is None:
        subtotal_indices = []
        
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        wb = writer.book
        ws = writer.sheets['Sheet1']
        
        # 스타일 정의
        header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        header_font = Font(name='맑은 고딕', bold=True, color="000000")
        subtotal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        subtotal_font = Font(name='맑은 고딕', bold=True)
        zebra_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        data_font = Font(name='맑은 고딕')
        
        center_align = Alignment(horizontal='center', vertical='center')
        
        vertical_border = Border(
            top=Side(border_style=None), 
            bottom=Side(border_style=None),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        double_border = Border(
            top=Side(style='double'), 
            bottom=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        bottom_double_border = Border(
            top=Side(border_style=None),
            bottom=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # 헤더 스타일 (Top 1 Row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
             
        ws.freeze_panes = 'A2'
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 데이터 서식
        for row in range(2, max_row + 1):
            is_subtotal = (row - 2) in subtotal_indices
            is_last_row = (row == max_row)
            is_even = (row % 2 == 0)
            
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # 공통 설정
                cell.font = data_font
                cell.alignment = center_align
                
                # 조건부 테두리 및 배경색
                current_border = vertical_border
                
                if is_even and not is_subtotal:
                    cell.fill = zebra_fill
                    
                if is_subtotal:
                    cell.fill = subtotal_fill
                    cell.font = subtotal_font
                    current_border = double_border
                elif is_last_row:
                    current_border = bottom_double_border
                
                cell.border = current_border
                
                # 숫자 및 날짜 서식
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                if isinstance(cell.value, datetime) or isinstance(cell.value, pd.Timestamp):
                    cell.number_format = 'yyyy-mm-dd'
                    
        # 열 너비 자동 맞춤
        for col in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            length = sum(1.5 if ord(c) > 127 else 1 for c in line)
                            if length > max_length:
                                max_length = length
                except:
                    pass
            adjusted_width = max_length * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width + 2, 50)
            
    return output.getvalue()

def add_subtotal_rows(df, group_col):
    new_rows = []
    subtotal_indices = []
    current_index = 0
    
    for name, group in df.groupby(group_col, sort=False):
        for _, row in group.iterrows():
            new_rows.append(row.to_dict())
            current_index += 1
            
        subtotal = {}
        for col in df.columns:
            if col == group_col:
                subtotal[col] = f"[{name} 소계]"
            elif col in ["Contract_ID", "contract_id", "id", "ID", "연도", "월", "계약일", "계약시작일", "계약종료일"]:
                subtotal[col] = None
            elif pd.api.types.is_numeric_dtype(df[col]):
                subtotal[col] = group[col].sum()
            else:
                subtotal[col] = None
        new_rows.append(subtotal)
        subtotal_indices.append(current_index)
        current_index += 1
        
    return pd.DataFrame(new_rows), subtotal_indices

st.set_page_config(
    page_title="상업용 부동산 자산관리 (PM/AM)",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS for Premium Design (cached)
@st.cache_resource
def get_custom_css():
    return """
<style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+KR:wght@400;500;700&display=swap');
    
    html, body, .stApp {
        font-family: 'Pretendard', 'Inter', 'Noto Sans KR', sans-serif;
        -webkit-font-smoothing: antialiased;
    }
    
    .block-container {
        padding-top: 2rem !important;
    }
    
    .stApp {
        background-color: #F8FAFC;
    }
    
    html, body, p, span, div {
        color: #334155;
    }

    h1, h2, h3 {
        color: #1e293b;
        font-weight: 700 !important;
        letter-spacing: -0.5px;
    }
    
    h1 {
        padding-bottom: 1rem;
        border-bottom: 2px solid #e2e8f0;
        margin-bottom: 2rem;
    }
    
    /* Buttons */
    .stButton>button {
        background: #ffffff;
        color: #334155 !important;
        border: 1px solid #CBD5E1;
        border-radius: 6px;
        padding: 0.4rem 1rem;
        font-weight: 500;
        font-size: 0.9rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
    }
    
    .stButton>button:hover {
        border-color: #005EB8;
        color: #005EB8 !important;
        background: #F8FAFC;
    }

    /* Cards and Containers */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border: 1px solid #E2E8F0 !important;
        border-radius: 8px !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
    }
    
    [data-testid="metric-container"] {
        background-color: #ffffff;
        border: 1px solid #E2E8F0;
        border-radius: 8px;
        padding: 1.2rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    [data-testid="stMetricValue"] > div {
        color: #005EB8 !important;
        font-weight: 800 !important;
    }
    [data-testid="stMetricLabel"] > div {
        color: #64748b !important;
    }
    
    .stAlert {
        border-radius: 8px;
        border: none;
        background-color: #ffffff;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background-color: transparent;
        padding: 0;
        border-bottom: 1px solid #E2E8F0;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 48px;
        padding-left: 16px;
        padding-right: 16px;
        white-space: nowrap !important;
        background-color: transparent;
        border-radius: 0;
        color: #64748B;
        font-weight: 500;
        font-size: 14px !important;
        border-bottom: 2px solid transparent;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        color: #005EB8 !important;
        font-weight: 600;
        border-bottom: 2px solid #005EB8 !important;
    }

    /* Forms & Inputs */
    .stTextInput>div>div>input, 
    .stNumberInput>div>div>input, 
    .stSelectbox>div>div>div,
    .stTextArea>div>div>textarea {
        border: 1px solid #CBD5E1 !important;
        border-radius: 6px;
        transition: all 0.3s ease;
        box-shadow: none !important;
    }
    .stTextInput>div>div>input:focus, 
    .stNumberInput>div>div>input:focus, 
    .stSelectbox>div>div>div:focus,
    .stTextArea>div>div>textarea:focus {
        border-color: #005EB8 !important;
        box-shadow: 0 0 0 1px #005EB8 !important;
    }

.custom-st-table th:nth-child(1), .custom-st-table td:nth-child(1) {{
    position: -webkit-sticky;
    position: sticky;
    left: 0;
    z-index: 5;
    background-color: white;
    border-right: none !important;
}}
.custom-st-table th:nth-child(1) {
    z-index: 15;
    background-color: #F8FAFC !important;
    color: #334155 !important;
}
.custom-st-table tr:hover td:nth-child(1) {
    background-color: #f8fafc;
}

    /* -------------------------------------------------------------------------- */
    /* UI/UX 레이아웃 교정 (픽셀 매칭 및 줄바꿈 방지)                           */
    /* -------------------------------------------------------------------------- */
    
    /* 1. 절대 줄바꿈 금지 (No-Wrap 강제) */
    .stButton>button, 
    [data-baseweb="tab"], 
    td, th, 
    [data-testid="stMetricLabel"], 
    [data-testid="stMetricValue"] {
        white-space: nowrap !important;
        word-break: keep-all !important;
        text-overflow: ellipsis;
    }

    /* 2. 입력창 및 버튼 규격(높이) 완벽 통일 */
    .stTextInput input, 
    .stNumberInput input, 
    .stSelectbox div[data-baseweb="select"], 
    .stButton button {
        height: 42px !important;
        line-height: 42px !important;
        margin: 0 !important;
    }

    /* 3. 대시보드 메트릭 박스(KPI) 동일 높이화 */
    [data-testid="metric-container"] {
        min-height: 130px !important;
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
    }
    /* -------------------------------------------------------------------------- */

</style>
""",
    unsafe_allow_html=True,
)

st.markdown(get_custom_css(), unsafe_allow_html=True)


# Helper function to send email with attachment
def send_email_with_attachment(to_email, subject, body, file_bytes, file_name, mime_type):
    if "email" not in st.secrets:
        return False, "Streamlit Cloud의 [Settings] -> [Secrets] 에 [email] 계정 정보를 입력해주세요!"
        
    try:
        user = st.secrets["email"]["user"]
        password = st.secrets["email"]["password"]
        host = st.secrets.get("email", {}).get("host", "smtp.gmail.com")
        port = st.secrets.get("email", {}).get("port", 465)
        
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = user
        msg["To"] = to_email
        msg.set_content(body)
        
        maintype, subtype = mime_type.split("/", 1)
        msg.add_attachment(file_bytes, maintype=maintype, subtype=subtype, filename=file_name)
        
        if port == 465:
            with smtplib.SMTP_SSL(host, port) as server:
                server.login(user, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port) as server:
                server.starttls()
                server.login(user, password)
                server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)



# Constants for currency conversion (should be updated via API in production)
CURRENCY_RATES = {
    "USD_TO_KRW": 1400.0,
    "PY_TO_SQM": 3.3058,
    "PY_TO_SF": 35.5832
}

# Performance optimization: Vectorized rent calculation
@st.cache_data(ttl=1800, show_spinner=False)
def calculate_rent_for_period_optimized(start_date, end_date, initial_rent, initial_maint, rent_schedule_json, rf_details, currency, year, month):
    """
    Optimized rent calculation for a specific month - calculates once and caches
    """
    import json
    _, last_day = calendar.monthrange(year, month)
    curr_month_start = datetime(year, month, 1)
    curr_month_end = datetime(year, month, last_day)
    
    overlap_start = max(start_date, curr_month_start)
    overlap_end = min(end_date, curr_month_end)
    
    if overlap_start > overlap_end:
        return 0.0, 0.0
    
    # Parse schedule once
    schedule = []
    if rent_schedule_json:
        try:
            schedule = json.loads(rent_schedule_json)
            schedule.sort(key=lambda x: pd.to_datetime(x["start_date"]))
        except:
            pass
    
    # Calculate total days in overlap period
    total_days = (overlap_end - overlap_start).days + 1
    
    # Check if entire month is rent-free
    month_str = f"{year}-{month:02d}"
    is_rf_month = month_str in rf_details if rf_details else False
    
    if is_rf_month:
        return 0.0, initial_maint
    
    # Find applicable rate for this month
    applicable_rent = initial_rent
    applicable_maint = initial_maint
    
    if schedule:
        for period in schedule:
            s_date = pd.to_datetime(period["start_date"]).date()
            e_date = pd.to_datetime(period["end_date"]).date()
            
            if s_date <= curr_month_start <= e_date or s_date <= curr_month_end <= e_date:
                applicable_rent = float(period.get("rent", initial_rent))
                applicable_maint = float(period.get("maint", initial_maint))
                break
    
    # Calculate pro-rated amounts
    rent_total = (applicable_rent / last_day) * total_days
    maint_total = (applicable_maint / last_day) * total_days
    
    # Round based on currency
    if currency == "KRW":
        rent_total = int(rent_total // 10) * 10
        maint_total = int(maint_total // 10) * 10
    else:
        rent_total = round(rent_total, 2)
        maint_total = round(maint_total, 2)
    
    return rent_total, maint_total

def get_scheduled_amount(rent_schedule_json, target_date, default_rent, default_maint, currency="KRW"):
    import json
    if rent_schedule_json:
        try:
            schedule = json.loads(rent_schedule_json)
            schedule.sort(key=lambda x: pd.to_datetime(x["start_date"]))
            
            last_known_rent = default_rent
            last_known_maint = default_maint
            
            for period in schedule:
                s_date = pd.to_datetime(period["start_date"]).date()
                e_date = pd.to_datetime(period["end_date"]).date()
                t_date = target_date.date()
                
                if t_date < s_date and last_known_rent == default_rent:
                    return default_rent, default_maint
                    
                if s_date <= t_date <= e_date:
                    return float(period.get("rent", 0.0)), float(period.get("maint", 0.0))
                
                if t_date > e_date:
                    last_known_rent = float(period.get("rent", 0.0))
                    last_known_maint = float(period.get("maint", 0.0))
            
            return last_known_rent, last_known_maint
        except:
            pass
    return default_rent, default_maint


def highlight_total_row(row):
    is_total = False
    for val in row:
        if str(val).strip() == "TOTAL":
            is_total = True
            break
    if is_total:
        return ["background-color: #F8FAFC; font-weight: bold; color: #0F172A"] * len(
            row
        )
    return [""] * len(row)


# Helper function to center align dataframe and style headers/totals
def center_styler(df):
    styler = df.style.apply(highlight_total_row, axis=1)
    return styler


# Helper function to display styled table as HTML with scrolling
def display_styled_table(df, freeze_cols=1, format_dict=None, custom_css=""):
    import uuid
    import streamlit.components.v1 as components

    if hasattr(df, "data"):
        df = df.data

    def format_money(x):
        if pd.isna(x):
            return ""
        try:
            val = float(x)
            if abs(val) >= 1_000_000:
                return f"{val/1_000_000:,.1f}백만"
            return f"{val:,.0f}"
        except:
            return str(x)

    auto_format = {}
    for col in df.columns:
        col_str = str(col)
        if any(k in col_str for k in ["면적", "비율", "율", "비중", "수익률"]):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = "{:,.2f}"
        elif any(
            k in col_str
            for k in [
                "금액",
                "보증금",
                "임대료",
                "관리비",
                "수익",
                "비용",
                "단가",
                "NOC",
                "월임대료",
                "월관리비",
                "합계",
            ]
        ):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = format_money

    if format_dict:
        auto_format.update(format_dict)

    styler = df.style.apply(highlight_total_row, axis=1)
    if auto_format:
        styler = styler.format(auto_format)
    try:
        styler = styler.hide(axis="index")
    except Exception:
        try:
            styler = styler.hide_index()
        except:
            pass

    uid = "tbl_" + uuid.uuid4().hex[:8]
    html = styler.to_html()

    if freeze_cols == 4:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; min-width: 150px; max-width: 150px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}
        .{uid} th:nth-child(2), .{uid} td:nth-child(2) {{ position: -webkit-sticky; position: sticky; left: 150px; z-index: 5; min-width: 100px; max-width: 100px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}
        .{uid} th:nth-child(3), .{uid} td:nth-child(3) {{ position: -webkit-sticky; position: sticky; left: 250px; z-index: 5; min-width: 200px; max-width: 200px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; white-space: normal; word-break: break-all; }}
        .{uid} th:nth-child(4), .{uid} td:nth-child(4) {{ position: -webkit-sticky; position: sticky; left: 450px; z-index: 5; min-width: 60px; max-width: 60px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}

        .{uid} th:nth-child(-n+4) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; text-align: center !important; }}
        """
    else:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; border-right: 1px solid #E2E8F0 !important; }}
        .{uid} th:nth-child(1) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; }}
        """

    wrapper = f"""
<html>
<head>
<style>
body {{ margin: 0; font-family: 'Pretendard', 'Inter', sans-serif; -webkit-font-smoothing: antialiased; }}
.custom-st-table.{uid} {{
    width: 100%;
    max-width: 100%;
    overflow-x: auto;
    max-height: 550px;
    overflow-y: auto;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}}
.custom-st-table.{uid} table {{
    width: 100%;
    border-collapse: collapse;
    background-color: white;
    font-size: 0.95rem;
    white-space: nowrap;
}}
.custom-st-table.{uid} th {{
    background-color: #F8FAFC !important;
    color: #334155 !important;
    font-weight: 600 !important;
    text-align: center !important;
    padding: 0.75rem 1rem !important;
    border-bottom: 1px solid #E2E8F0 !important;
    border-right: 1px solid #E2E8F0 !important;
    position: -webkit-sticky;
    position: sticky;
    top: 0;
    z-index: 10;
}}
.custom-st-table.{uid} td {{
    padding: 0.75rem 1rem !important;
    text-align: right !important;
    border-bottom: 1px solid #E2E8F0 !important;
    border-right: 1px solid #E2E8F0 !important;
    color: #334155;
}}
.custom-st-table.{uid} td:first-child {{
    text-align: left !important;
}}
.custom-st-table.{uid} td:last-child, .custom-st-table.{uid} th:last-child {{
    border-right: none !important;
}}
.custom-st-table.{uid} tr:nth-child(odd) td {{
    background-color: #ffffff;
}}
.custom-st-table.{uid} tr:nth-child(even) td {{
    background-color: #F8F9FA;
}}
.custom-st-table.{uid} tr:nth-child(odd) td:nth-child(even) {{
    background-color: #F1F3F5;
}}
.custom-st-table.{uid} tr:nth-child(even) td:nth-child(even) {{
    background-color: #E9ECEF;
}}
.custom-st-table.{uid} tr:hover td {{
    background-color: #E2E8F0 !important;
}}
{freeze_css}
{custom_css.replace('{uid}', uid)}
</style>
</head>
<body>
<div class="custom-st-table {uid}">
    {html}
</div>
</body>
</html>
"""
    components.html(wrapper, height=560, scrolling=False)


# DB Init
@st.cache_resource
def init_db():
    db_url = st.secrets["DATABASE_URL"]
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS Asset_Area (
            asset_name TEXT,
            floor TEXT,
            exclusive_area REAL,
            common_area REAL,
            total_area REAL,
            PRIMARY KEY (asset_name, floor)
        )
    """)
    conn.commit()
    try:
        c.execute("ALTER TABLE Asset_Area ADD COLUMN bank_area REAL DEFAULT 0.0")
        conn.commit()
    except psycopg2.DatabaseError:
        conn.rollback()
        pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS Lease_Contracts (
            contract_id SERIAL PRIMARY KEY,
            asset_name TEXT,
            floor TEXT,
            company_name TEXT,
            contract_date DATE,
            start_date DATE,
            end_date DATE,
            contract_area REAL,
            deposit REAL,
            monthly_rent REAL,
            monthly_maintenance_fee REAL,
            total_rent_free_months INTEGER,
            rent_free_details TEXT
        )
    """)

    new_columns = [
        ("status", "TEXT DEFAULT 'ACTIVE'"),
        ("deposit_return_date", "DATE"),
        ("penalty_yn", "TEXT"),
        ("penalty_amount", "REAL"),
        ("parent_contract_id", "INTEGER"),
        ("currency", "TEXT DEFAULT 'KRW'"),
        ("floor_details", "TEXT"),
        ("escalation_cycle_years", "INTEGER"),
        ("rent_inc_rate", "REAL"),
        ("maint_inc_rate", "REAL"),
        ("contract_exclusive_area", "REAL"),
        ("rent_schedule", "TEXT"),
        ("remarks", "TEXT"),
    ]
    for col_name, col_type in new_columns:
        try:
            c.execute(f"ALTER TABLE Lease_Contracts ADD COLUMN {col_name} {col_type}")
            if col_name == "status":
                c.execute(
                    "UPDATE Lease_Contracts SET status = 'ACTIVE' WHERE status IS NULL"
                )
            conn.commit()
        except psycopg2.DatabaseError:
            conn.rollback()
            pass

    conn.commit()

    c.execute("DROP TABLE IF EXISTS RentRoll_Overrides")
    c.execute("""
        CREATE TABLE IF NOT EXISTS RentRoll_Overrides (
            contract_id INTEGER,
            floor TEXT,
            year INTEGER,
            month INTEGER,
            over_rent REAL,
            over_maint REAL,
            PRIMARY KEY (contract_id, floor, year, month)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS Contract_History (
            history_id SERIAL PRIMARY KEY,
            contract_id INTEGER,
            action_type TEXT,
            action_date DATE,
            action_month TEXT,
            details TEXT
        )
    """)
    conn.commit()
    return conn


conn = init_db()


@st.cache_resource
def get_engine():
    from sqlalchemy import create_engine
    return create_engine(
        st.secrets["DATABASE_URL"], pool_size=5, max_overflow=2, pool_recycle=300, pool_pre_ping=True
    )


engine = get_engine()


@st.cache_data(ttl=600, show_spinner=False)
def fetch_data(query, _eng=None):
    if _eng is None:
        _eng = get_engine()
    return pd.read_sql(query, _eng)


def get_floor_sort_key(floor_str):
    if not isinstance(floor_str, str):
        return -9999
    f = floor_str.upper().replace("F", "").strip()
    if f.startswith("B"):
        try:
            return -int(f[1:])
        except:
            return -9999
    else:
        try:
            return int(f)
        except:
            return 0


def sort_df_by_asset_and_floor(df, asset_col="asset_name", floor_col="floor"):
    import pandas as pd

    asset_df = fetch_data("SELECT asset_name, total_area FROM Asset_Area")
    if not asset_df.empty:
        sorted_assets = (
            asset_df.groupby("asset_name")["total_area"]
            .sum()
            .sort_values(ascending=False)
            .index.tolist()
        )
    else:
        sorted_assets = []

    if asset_col in df.columns:
        df[asset_col] = pd.Categorical(
            df[asset_col], categories=sorted_assets, ordered=True
        )
        sort_cols = [asset_col]
        asc = [True]
    else:
        sort_cols = []
        asc = []

    if floor_col in df.columns:
        df["_floor_sort"] = df[floor_col].apply(get_floor_sort_key)
        sort_cols.append("_floor_sort")
        asc.append(False)

    if sort_cols:
        df = df.sort_values(sort_cols, ascending=asc)
        if "_floor_sort" in df.columns:
            df = df.drop(columns=["_floor_sort"])

    return df


st.title("🏢 상업용 부동산 자산관리 시스템 (PM/AM)")

with st.sidebar:
    st.header("🔔 D-180 만기 도래 알림 데스크")
    df_active = fetch_data(
        "SELECT asset_name, floor, company_name, end_date FROM Lease_Contracts WHERE status = 'ACTIVE'"
    )

    if not df_active.empty:
        df_active["end_date"] = pd.to_datetime(df_active["end_date"], errors="coerce")
        today = pd.to_datetime(datetime.now().date())
        df_active["d_day"] = (df_active["end_date"] - today).dt.days

        df_expiring = df_active[
            (df_active["d_day"] <= 180) & (df_active["d_day"] >= 0)
        ].sort_values("d_day")

        if not df_expiring.empty:
            for _, row in df_expiring.iterrows():
                if row["d_day"] <= 30:
                    st.error(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
                else:
                    st.warning(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
        else:
            st.info("현재 6개월 내 만기 도래 계약이 없습니다.")
    else:
        st.info("현재 활성 계약이 없습니다.")


(
    tab_master_dashboard,
    tab_market_research,
    tab_asset_view,
    tab_stacking_plan,
    tab_lease_info,
    tab_rent_roll,
    tab_asset_update,
    tab_contract_update,
    tab_history,
) = st.tabs(
    [
        "🌐 마스터 대시보드",
        "📈 시장 동향 리서치",
        "📊 자산별 면적 현황",
        "🏢 스태킹 플랜",
        "📝 자산별 임대정보",
        "💰 렌트롤 (Rent Roll)",
        "✏️ 자산정보 업데이트",
        "✍️ 계약 업데이트",
        "🕒 업데이트 이력 관리",
    ]
)


def get_months_between(start_date, end_date):
    months = []
    if not start_date or not end_date or start_date > end_date:
        return months
    current = datetime(start_date.year, start_date.month, 1)
    end = datetime(end_date.year, end_date.month, 1)
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return months



@st.cache_data(ttl=86400)
def fetch_market_research_data():
    import random
    import pandas as pd
    
    regions = ["서울", "경기", "인천", "부산", "대구", "광주", "대전"]
    sub_regions = {
        "서울": ["강남대로", "테헤란로", "도산대로", "여의도", "광화문", "명동", "홍대합정"],
        "경기": ["분당", "판교", "일산", "평촌"],
        "인천": ["부평", "구월", "송도"],
        "부산": ["서면", "해운대", "광복동"],
        "대구": ["동성로", "수성구"],
        "광주": ["상무지구", "충장로"],
        "대전": ["둔산", "은행동"]
    }
    asset_types = ["오피스", "소규모 상가", "중대형 상가"]
    quarters = ["2023 1Q", "2023 2Q", "2023 3Q", "2023 4Q", "2024 1Q", "2024 2Q"]
    
    data = []
    for r in regions:
        for sr in sub_regions[r]:
            for at in asset_types:
                for q in quarters:
                    base_rent = random.uniform(15000, 35000) if at == "오피스" else random.uniform(20000, 60000)
                    if r == "서울":
                        base_rent *= 1.5
                    vacancy = random.uniform(2.0, 15.0)
                    data.append({
                        "지역명(시/도)": r,
                        "세부 상권명": sr,
                        "자산 유형": at,
                        "기준 분기": q,
                        "㎡당 임대료": round(base_rent),
                        "공실률(%)": round(vacancy, 1)
                    })
    
    df = pd.DataFrame(data)
    df["평당 임대료"] = (df["㎡당 임대료"] * 3.3058).round().astype(int)
    return df

# ==========================================
# Tab 0: 마스터 대시보드
# ==========================================
with tab_master_dashboard:
    st.header("🌐 마스터 대시보드 (Executive Summary)")
    import plotly.express as px
    import numpy as np

    df_assets_md = fetch_data("SELECT * FROM Asset_Area")
    df_leases_md = fetch_data("SELECT * FROM Lease_Contracts WHERE status = 'ACTIVE'")

    if df_assets_md.empty:
        st.warning("등록된 자산 정보가 없습니다.")
    else:
        # [1단: 직관적 핵심 지표 (KPIs)]
        total_assets = df_assets_md["asset_name"].nunique()
        total_exclusive_area = df_assets_md["exclusive_area"].sum()
        total_bank_area = df_assets_md["bank_area"].sum() if "bank_area" in df_assets_md.columns else 0
        
        today_md = datetime.now()
        
        if not df_leases_md.empty:
            df_leases_md["start_date"] = pd.to_datetime(df_leases_md["start_date"])
            df_leases_md["end_date"] = pd.to_datetime(df_leases_md["end_date"])
            active_leases = df_leases_md[
                (df_leases_md["start_date"] <= today_md)
                & (df_leases_md["end_date"] >= today_md)
            ].copy()
        else:
            active_leases = pd.DataFrame()
            
        total_active_area = active_leases["contract_exclusive_area"].sum() if not active_leases.empty else 0
        
        # 1. 통합 관리 자산 및 임대율
        total_occupied = total_active_area + total_bank_area
        occupancy_rate = (total_occupied / total_exclusive_area * 100) if total_exclusive_area > 0 else 0
        
        # 2. 전체 임차사 및 활성 계약 수
        unique_companies = active_leases["company_name"].nunique() if not active_leases.empty else 0
        total_contracts = len(active_leases)
        
        # 3. 당월 총 청구 수익
        if not active_leases.empty:
            def get_krw_val(row, col):
                val = row[col] if pd.notnull(row[col]) else 0
                if row.get("currency", "KRW") == "USD":
                    return val * CURRENCY_RATES["USD_TO_KRW"]
                return val

            active_leases["rent_krw"] = active_leases.apply(lambda r: get_krw_val(r, "monthly_rent"), axis=1)
            active_leases["maint_krw"] = active_leases.apply(lambda r: get_krw_val(r, "monthly_maintenance_fee"), axis=1)
            monthly_revenue = active_leases["rent_krw"].sum() + active_leases["maint_krw"].sum()
        else:
            monthly_revenue = 0

        # 4. 통합 공실 면적
        vacant_area = max(0, total_exclusive_area - total_occupied)

        st.markdown("### 📊 포트폴리오 핵심 지표 (Executive KPIs)")
        c1, c2, c3, c4 = st.columns(4, vertical_alignment="bottom")
        c1.metric(
            "통합 관리 자산 및 임대율",
            f"{total_assets}개",
            f"임대율: {occupancy_rate:.1f}%",
        )
        c2.metric(
            "전체 임차사 및 활성 계약 수",
            f"{unique_companies}개 사",
            f"계약 {total_contracts}건",
        )
        mr_str = f"₩ {monthly_revenue/1000000:,.0f}백만" if monthly_revenue >= 1000000 else f"₩ {monthly_revenue:,.0f}"
        c3.metric(
            "당월 총 청구 수익",
            mr_str,
        )
        c4.metric(
            "통합 공실 면적",
            f"{vacant_area:,.1f} 평",
        )

        st.markdown("---")

        # [2단: 3대 핵심 시각화 차트]
        st.markdown("### 📈 통합 데이터 시각화 (Portfolio Analytics)")
        p1, p2 = st.columns(2, vertical_alignment="bottom")
        
        sc_palette = ["#005EB8", "#00A546", "#38BDF8", "#34D399", "#94A3B8"]
        
        with p1:
            with st.container(border=True):
                st.markdown("#### 포트폴리오 공간 점유 현황")
                pie_df = pd.DataFrame({
                    "Category": ["은행/지점 사용 면적", "일반 테넌트 임대 면적", "공실 면적"],
                    "Area": [total_bank_area, total_active_area, vacant_area]
                })
                pie_df = pie_df[pie_df["Area"] > 0]
                if not pie_df.empty:
                    fig_pie = px.pie(
                        pie_df, 
                        names="Category", 
                        values="Area", 
                        hole=0.4,
                        color_discrete_sequence=sc_palette
                    )
                    fig_pie.update_layout(
                        margin=dict(l=0, r=0, t=30, b=0),
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
                    )
                    fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                    st.plotly_chart(fig_pie, use_container_width=True)
                else:
                    st.info("데이터 없음")

        with p2:
            with st.container(border=True):
                st.markdown("#### 연도별 만기 도래 면적 (향후 5년)")
                if not active_leases.empty:
                    current_year = today_md.year
                    active_leases["end_year"] = active_leases["end_date"].dt.year
                    df_exp = active_leases[
                        (active_leases["end_year"] >= current_year) & 
                        (active_leases["end_year"] <= current_year + 5)
                    ]
                    if not df_exp.empty:
                        df_exp_grp = df_exp.groupby("end_year")["contract_area"].sum().reset_index()
                        df_exp_grp["end_year"] = df_exp_grp["end_year"].astype(str)
                        fig_exp = px.bar(
                            df_exp_grp,
                            x="end_year",
                            y="contract_area",
                            color_discrete_sequence=[sc_palette[0]]
                        )
                        fig_exp.update_layout(
                            xaxis_title="만기 연도",
                            yaxis_title="만기 도래 면적(평)",
                            margin=dict(l=0, r=0, t=30, b=0),
                            plot_bgcolor='rgba(0,0,0,0)',
                            paper_bgcolor='rgba(0,0,0,0)',
                            xaxis=dict(showgrid=False, zeroline=False),
                            yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False),
                            showlegend=False
                        )
                        fig_exp.update_traces(texttemplate='%{y:,.0f}', textposition='outside')
                        st.plotly_chart(fig_exp, use_container_width=True)
                    else:
                        st.info("향후 5년 내 만기 도래 계약 없음")
                else:
                    st.info("데이터 없음")
                    
        st.markdown("---")
        
        # [3단: 자산별 평단가 비교]
        with st.container(border=True):
            st.markdown("#### 자산별 평균 평단가 비교 (임대료 및 관리비)")
            if not active_leases.empty:
                valid_leases = active_leases[active_leases["contract_area"].astype(float) > 0].copy()
                valid_leases["rent_unit"] = valid_leases["rent_krw"] / valid_leases["contract_area"].astype(float)
                valid_leases["maint_unit"] = valid_leases["maint_krw"] / valid_leases["contract_area"].astype(float)
                df_unit = valid_leases.groupby("asset_name")[["rent_unit", "maint_unit"]].mean().reset_index()
                
                df_unit["total_unit"] = df_unit["rent_unit"] + df_unit["maint_unit"]
                df_unit = df_unit.sort_values(by="total_unit", ascending=False)
                
                df_melt = df_unit.melt(
                    id_vars=["asset_name"], 
                    value_vars=["rent_unit", "maint_unit"], 
                    var_name="Type", 
                    value_name="Unit_Price"
                )
                df_melt["Type"] = df_melt["Type"].replace({"rent_unit": "임대료", "maint_unit": "관리비"})
                
                fig_bar2 = px.bar(
                    df_melt,
                    x="asset_name",
                    y="Unit_Price",
                    color="Type",
                    barmode="group",
                    color_discrete_sequence=["#005EB8", "#34D399"]
                )
                fig_bar2.update_layout(
                    xaxis_title="자산명",
                    yaxis_title="평단가 (KRW)",
                    margin=dict(l=0, r=0, t=30, b=0),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=False, zeroline=False),
                    yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False),
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                fig_bar2.update_traces(texttemplate='₩ %{y:,.0f}', textposition='outside')
                st.plotly_chart(fig_bar2, use_container_width=True)
            else:
                st.info("데이터 없음")