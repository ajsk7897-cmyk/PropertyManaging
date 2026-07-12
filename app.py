import streamlit as st
import pandas as pd
import plotly.express as px
import psycopg2
import psycopg2.extras
import calendar
from datetime import datetime, timedelta
import json
import io
import smtplib
from email.message import EmailMessage
import os
from openpyxl import Workbook, load_workbook

st.set_page_config(
    page_title="상업용 부동산 자산관리 (PM/AM)",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS for Premium Design
st.markdown(
    """
<style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+KR:wght@400;500;700&display=swap');
    
    html, body, .stApp {
        font-family: 'Pretendard', 'Inter', 'Noto Sans KR', sans-serif;
        -webkit-font-smoothing: antialiased;
    }
    
    .block-container {
        padding-top: 2rem !important;
    }
    
    .stApp {
        background-color: #F8FAFC;
    }
    
    html, body, p, span, div {
        color: #334155;
    }

    h1, h2, h3 {
        color: #1e293b;
        font-weight: 700 !important;
        letter-spacing: -0.5px;
    }
    
    h1 {
        padding-bottom: 1rem;
        border-bottom: 2px solid #e2e8f0;
        margin-bottom: 2rem;
    }
    
    /* Buttons */
    .stButton>button {
        background: #ffffff;
        color: #334155 !important;
        border: 1px solid #CBD5E1;
        border-radius: 6px;
        padding: 0.4rem 1rem;
        font-weight: 500;
        font-size: 0.9rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
    }
    
    .stButton>button:hover {
        border-color: #005EB8;
        color: #005EB8 !important;
        background: #F8FAFC;
    }

    /* Cards and Containers */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border: 1px solid #E2E8F0 !important;
        border-radius: 8px !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
    }
    
    [data-testid="metric-container"] {
        background-color: #ffffff;
        border: 1px solid #E2E8F0;
        border-radius: 8px;
        padding: 1.2rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    [data-testid="stMetricValue"] > div {
        color: #005EB8 !important;
        font-weight: 800 !important;
    }
    [data-testid="stMetricLabel"] > div {
        color: #64748b !important;
    }
    
    .stAlert {
        border-radius: 8px;
        border: none;
        background-color: #ffffff;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background-color: transparent;
        padding: 0;
        border-bottom: 1px solid #E2E8F0;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 48px;
        padding-left: 16px;
        padding-right: 16px;
        white-space: nowrap !important;
        background-color: transparent;
        border-radius: 0;
        color: #64748B;
        font-weight: 500;
        font-size: 14px !important;
        border-bottom: 2px solid transparent;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        color: #005EB8 !important;
        font-weight: 600;
        border-bottom: 2px solid #005EB8 !important;
    }

    /* Forms & Inputs */
    .stTextInput>div>div>input, 
    .stNumberInput>div>div>input, 
    .stSelectbox>div>div>div,
    .stTextArea>div>div>textarea {
        border: 1px solid #CBD5E1 !important;
        border-radius: 6px;
        transition: all 0.3s ease;
        box-shadow: none !important;
    }
    .stTextInput>div>div>input:focus, 
    .stNumberInput>div>div>input:focus, 
    .stSelectbox>div>div>div:focus,
    .stTextArea>div>div>textarea:focus {
        border-color: #005EB8 !important;
        box-shadow: 0 0 0 1px #005EB8 !important;
    }

.custom-st-table th:nth-child(1), .custom-st-table td:nth-child(1) {{
    position: -webkit-sticky;
    position: sticky;
    left: 0;
    z-index: 5;
    background-color: white;
    border-right: none !important;
}}
.custom-st-table th:nth-child(1) {
    z-index: 15;
    background-color: #F8FAFC !important;
    color: #334155 !important;
}
.custom-st-table tr:hover td:nth-child(1) {
    background-color: #f8fafc;
}
</style>
""",
    unsafe_allow_html=True,
)


# Helper function to send email with attachment
def send_email_with_attachment(to_email, subject, body, file_bytes, file_name, mime_type):
    try:
        user = st.secrets["email"]["user"]
        password = st.secrets["email"]["password"]
        host = st.secrets.get("email", {}).get("host", "smtp.gmail.com")
        port = st.secrets.get("email", {}).get("port", 465)
        
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = user
        msg["To"] = to_email
        msg.set_content(body)
        
        maintype, subtype = mime_type.split("/", 1)
        msg.add_attachment(file_bytes, maintype=maintype, subtype=subtype, filename=file_name)
        
        if port == 465:
            with smtplib.SMTP_SSL(host, port) as server:
                server.login(user, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port) as server:
                server.starttls()
                server.login(user, password)
                server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)


def calculate_escalated_amount(row, target_date):
    start_date = pd.to_datetime(row["start_date"])
    initial_rent = (
        float(row["monthly_rent"])
        if "monthly_rent" in row and pd.notnull(row["monthly_rent"])
        else 0.0
    )
    initial_maint = (
        float(row["monthly_maintenance_fee"])
        if "monthly_maintenance_fee" in row
        and pd.notnull(row["monthly_maintenance_fee"])
        else 0.0
    )

    esc_cycle = (
        int(row["escalation_cycle_years"])
        if "escalation_cycle_years" in row and pd.notnull(row["escalation_cycle_years"])
        else 0
    )
    rent_inc = (
        float(row["rent_inc_rate"])
        if "rent_inc_rate" in row and pd.notnull(row["rent_inc_rate"])
        else 0.0
    )
    maint_inc = (
        float(row["maint_inc_rate"])
        if "maint_inc_rate" in row and pd.notnull(row["maint_inc_rate"])
        else 0.0
    )

    SYSTEM_BASE_DATE = datetime(2026, 7, 1)

    if start_date < SYSTEM_BASE_DATE:
        base_elapsed = (2026 - start_date.year) - (1 if 7 < start_date.month else 0)
        base_cycles = base_elapsed // esc_cycle if esc_cycle > 0 else 0
    else:
        base_cycles = 0

    target_elapsed = (target_date.year - start_date.year) - (
        1 if target_date.month < start_date.month else 0
    )
    target_cycles = target_elapsed // esc_cycle if esc_cycle > 0 else 0

    cycles_to_apply = target_cycles - base_cycles

    escalated_rent = initial_rent * ((1 + rent_inc / 100.0) ** cycles_to_apply)
    escalated_maint = initial_maint * ((1 + maint_inc / 100.0) ** cycles_to_apply)

    currency = (
        row["currency"] if "currency" in row and pd.notnull(row["currency"]) else "KRW"
    )

    if currency == "KRW":
        escalated_rent = (escalated_rent // 10) * 10
        escalated_maint = (escalated_maint // 10) * 10

    return escalated_rent, escalated_maint


def highlight_total_row(row):
    is_total = False
    for val in row:
        if str(val).strip() == "TOTAL":
            is_total = True
            break
    if is_total:
        return ["background-color: #F8FAFC; font-weight: bold; color: #0F172A"] * len(
            row
        )
    return [""] * len(row)


# Helper function to center align dataframe and style headers/totals
def center_styler(df):
    # CSS is already handled globally in display_styled_table wrapper.
    # Removing set_properties and set_table_styles massively speeds up styler.to_html()
    styler = df.style.apply(highlight_total_row, axis=1)
    return styler


# Helper function to display styled table as HTML with scrolling
def display_styled_table(df, freeze_cols=1, format_dict=None, custom_css=""):
    import uuid
    import streamlit.components.v1 as components

    if hasattr(df, "data"):
        df = df.data

    def format_money(x):
        if pd.isna(x):
            return ""
        try:
            val = float(x)
            if abs(val) >= 1_000_000:
                return f"{val/1_000_000:,.1f}백만"
            return f"{val:,.0f}"
        except:
            return str(x)

    auto_format = {}
    for col in df.columns:
        col_str = str(col)
        if any(k in col_str for k in ["면적", "비율", "율", "비중", "수익률"]):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = "{:,.2f}"
        elif any(
            k in col_str
            for k in [
                "금액",
                "보증금",
                "임대료",
                "관리비",
                "수익",
                "비용",
                "단가",
                "NOC",
                "월임대료",
                "월관리비",
                "합계",
            ]
        ):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = format_money

    if format_dict:
        auto_format.update(format_dict)

    styler = df.style.apply(highlight_total_row, axis=1)
    if auto_format:
        styler = styler.format(auto_format)
    try:
        styler = styler.hide(axis="index")
    except Exception:
        try:
            styler = styler.hide_index()
        except:
            pass

    uid = "tbl_" + uuid.uuid4().hex[:8]
    html = styler.to_html()

    if freeze_cols == 4:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; min-width: 120px; max-width: 120px; border-right: none !important; }}
        .{uid} th:nth-child(2), .{uid} td:nth-child(2) {{ position: -webkit-sticky; position: sticky; left: 120px; z-index: 5; min-width: 80px; max-width: 80px; border-right: none !important; }}
        .{uid} th:nth-child(3), .{uid} td:nth-child(3) {{ position: -webkit-sticky; position: sticky; left: 200px; z-index: 5; min-width: 150px; max-width: 150px; border-right: none !important; }}
        .{uid} th:nth-child(4), .{uid} td:nth-child(4) {{ position: -webkit-sticky; position: sticky; left: 350px; z-index: 5; min-width: 60px; max-width: 60px; border-right: none !important; }}

        .{uid} th:nth-child(-n+4) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; }}
        """
    else:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; border-right: none !important; }}
        .{uid} th:nth-child(1) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; }}
        """

    wrapper = f"""
<html>
<head>
<style>
body {{ margin: 0; font-family: 'Pretendard', 'Inter', sans-serif; -webkit-font-smoothing: antialiased; }}
.custom-st-table.{uid} {{
    width: 100%;
    max-width: 100%;
    overflow-x: auto;
    max-height: 550px;
    overflow-y: auto;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}}
.custom-st-table.{uid} table {{
    width: 100%;
    border-collapse: collapse;
    background-color: white;
    font-size: 0.95rem;
    white-space: nowrap;
}}
.custom-st-table.{uid} th {{
    background-color: #F8FAFC !important;
    color: #334155 !important;
    font-weight: 600 !important;
    text-align: center !important;
    padding: 0.75rem 1rem !important;
    border-bottom: 1px solid #F1F5F9 !important;
    border-right: none !important;
    position: -webkit-sticky;
    position: sticky;
    top: 0;
    z-index: 10;
}}
.custom-st-table.{uid} td {{
    padding: 0.75rem 1rem !important;
    text-align: right !important;
    border-bottom: 1px solid #F1F5F9 !important;
    border-right: none !important;
    color: #334155;
}}
.custom-st-table.{uid} td:first-child {{
    text-align: left !important;
}}
.custom-st-table.{uid} th {{
    /* Removed vertical lines */
}}
.custom-st-table.{uid} td:last-child, .custom-st-table.{uid} th:last-child {{
    border-right: none;
}}
.custom-st-table.{uid} tr:nth-child(even) td {{
    background-color: #ffffff;
}}
.custom-st-table.{uid} tr:nth-child(odd) td {{
    background-color: #ffffff;
}}
.custom-st-table.{uid} tr:hover td {{
    background-color: #F8FAFC !important;
}}
{freeze_css}
{custom_css.replace('{uid}', uid)}
</style>
</head>
<body>
<div class="custom-st-table {uid}">
    {html}
</div>
</body>
</html>
"""
    components.html(wrapper, height=560, scrolling=False)


# DB Init
@st.cache_resource
def init_db():
    # Use Streamlit secrets for Supabase connection
    db_url = st.secrets["DATABASE_URL"]
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS Asset_Area (
            asset_name TEXT,
            floor TEXT,
            exclusive_area REAL,
            common_area REAL,
            total_area REAL,
            PRIMARY KEY (asset_name, floor)
        )
    """)
    try:
        c.execute("ALTER TABLE Asset_Area ADD COLUMN bank_area REAL DEFAULT 0.0")
    except psycopg2.DatabaseError:
        conn.rollback()
        pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS Lease_Contracts (
            contract_id SERIAL PRIMARY KEY,
            asset_name TEXT,
            floor TEXT,
            company_name TEXT,
            contract_date DATE,
            start_date DATE,
            end_date DATE,
            contract_area REAL,
            deposit REAL,
            monthly_rent REAL,
            monthly_maintenance_fee REAL,
            total_rent_free_months INTEGER,
            rent_free_details TEXT
        )
    """)

    # Alter Lease_Contracts to add new columns safely
    new_columns = [
        ("status", "TEXT DEFAULT 'ACTIVE'"),
        ("deposit_return_date", "DATE"),
        ("penalty_yn", "TEXT"),
        ("penalty_amount", "REAL"),
        ("parent_contract_id", "INTEGER"),
        ("currency", "TEXT DEFAULT 'KRW'"),
        ("floor_details", "TEXT"),
        ("escalation_cycle_years", "INTEGER"),
        ("rent_inc_rate", "REAL"),
        ("maint_inc_rate", "REAL"),
        ("contract_exclusive_area", "REAL"),
        ("remarks", "TEXT"),
    ]
    for col_name, col_type in new_columns:
        try:
            c.execute(f"ALTER TABLE Lease_Contracts ADD COLUMN {col_name} {col_type}")
            if col_name == "status":
                c.execute(
                    "UPDATE Lease_Contracts SET status = 'ACTIVE' WHERE status IS NULL"
                )
        except psycopg2.DatabaseError:
            conn.rollback()
            pass  # Column already exists

    c.execute("DROP TABLE IF EXISTS RentRoll_Overrides")
    c.execute("""
        CREATE TABLE IF NOT EXISTS RentRoll_Overrides (
            contract_id INTEGER,
            floor TEXT,
            year INTEGER,
            month INTEGER,
            over_rent REAL,
            over_maint REAL,
            PRIMARY KEY (contract_id, floor, year, month)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS Contract_History (
            history_id SERIAL PRIMARY KEY,
            contract_id INTEGER,
            action_type TEXT,
            action_date DATE,
            action_month TEXT,
            details TEXT
        )
    """)
    conn.commit()
    return conn


conn = init_db()


@st.cache_resource
def get_engine():
    from sqlalchemy import create_engine

    return create_engine(
        st.secrets["DATABASE_URL"], pool_size=5, max_overflow=2, pool_recycle=300, pool_pre_ping=True
    )


engine = get_engine()


@st.cache_data(ttl=600, show_spinner=False)
def fetch_data(query, _eng=None):
    if _eng is None:
        _eng = get_engine()
    return pd.read_sql(query, _eng)


def get_floor_sort_key(floor_str):
    if not isinstance(floor_str, str):
        return -9999
    f = floor_str.upper().replace("F", "").strip()
    if f.startswith("B"):
        try:
            return -int(f[1:])
        except:
            return -9999
    else:
        try:
            return int(f)
        except:
            return 0


def sort_df_by_asset_and_floor(df, asset_col="asset_name", floor_col="floor"):
    import pandas as pd

    asset_df = fetch_data("SELECT asset_name, total_area FROM Asset_Area")
    if not asset_df.empty:
        sorted_assets = (
            asset_df.groupby("asset_name")["total_area"]
            .sum()
            .sort_values(ascending=False)
            .index.tolist()
        )
    else:
        sorted_assets = []

    if asset_col in df.columns:
        df[asset_col] = pd.Categorical(
            df[asset_col], categories=sorted_assets, ordered=True
        )
        sort_cols = [asset_col]
        asc = [True]
    else:
        sort_cols = []
        asc = []

    if floor_col in df.columns:
        df["_floor_sort"] = df[floor_col].apply(get_floor_sort_key)
        sort_cols.append("_floor_sort")
        asc.append(False)

    if sort_cols:
        df = df.sort_values(sort_cols, ascending=asc)
        if "_floor_sort" in df.columns:
            df = df.drop(columns=["_floor_sort"])

    return df


st.title("🏢 상업용 부동산 자산관리 시스템 (PM/AM)")

with st.sidebar:
    st.header("🔔 D-180 만기 도래 알림 데스크")
    df_active = fetch_data(
        "SELECT asset_name, floor, company_name, end_date FROM Lease_Contracts WHERE status = 'ACTIVE'"
    )

    if not df_active.empty:
        df_active["end_date"] = pd.to_datetime(df_active["end_date"], errors="coerce")
        today = pd.to_datetime(datetime.now().date())
        df_active["d_day"] = (df_active["end_date"] - today).dt.days

        # Filter 180 days or less, and >= 0 (not expired yet, or we can show expired as well)
        df_expiring = df_active[
            (df_active["d_day"] <= 180) & (df_active["d_day"] >= 0)
        ].sort_values("d_day")

        if not df_expiring.empty:
            for _, row in df_expiring.iterrows():
                if row["d_day"] <= 30:
                    st.error(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
                else:
                    st.warning(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
        else:
            st.info("현재 6개월 내 만기 도래 계약이 없습니다.")
    else:
        st.info("현재 활성 계약이 없습니다.")


(
    tab_master_dashboard,
    tab_market_research,
    tab_asset_view,
    tab_stacking_plan,
    tab_lease_info,
    tab_rent_roll,
    tab_asset_update,
    tab_contract_update,
    tab_history,
) = st.tabs(
    [
        "🌐 마스터 대시보드",
        "📈 시장 동향 리서치",
        "📊 자산별 면적 현황",
        "🏢 스태킹 플랜",
        "📝 자산별 임대정보",
        "💰 렌트롤 (Rent Roll)",
        "✏️ 자산정보 업데이트",
        "✍️ 계약 업데이트",
        "🕒 업데이트 이력 관리",
    ]
)


def get_months_between(start_date, end_date):
    months = []
    if not start_date or not end_date or start_date > end_date:
        return months
    current = datetime(start_date.year, start_date.month, 1)
    end = datetime(end_date.year, end_date.month, 1)
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return months



@st.cache_data(ttl=86400)
def fetch_market_research_data():
    import random
    import pandas as pd
    import requests
    import streamlit as st
    
    # ---------------------------------------------------------
    # [실제 API 연동부] 
    # API 키는 .streamlit/secrets.toml의 R_ONE_API_KEY를 자동 참조합니다.
    # 추후 실제 Endpoint URL과 파라미터 구조가 확정되면 아래 주석을 풀고 연동합니다.
    # ---------------------------------------------------------
    api_key = st.secrets.get("R_ONE_API_KEY", None)
    api_endpoint = "https://api.reb.or.kr/v1/market/rent" # TODO: 실제 URL로 교체 필요
    
    if api_key and "TODO" not in api_endpoint:
        try:
            # 실제 연동 예시
            # response = requests.get(api_endpoint, params={"serviceKey": api_key, "format": "json"}, timeout=3)
            # response.raise_for_status()
            # items = response.json().get("response", {}).get("body", {}).get("items", [])
            # df = pd.DataFrame(items)
            # df["평당 임대료"] = (df["㎡당 임대료"] * 3.3058).round().astype(int)
            # return df
            pass
        except Exception as e:
            st.warning(f"API 연동 오류 (더미 데이터로 대체합니다): {e}")
            
    # ---------------------------------------------------------
    # API가 구성되지 않았거나 실패했을 때를 대비한 더미 데이터 생성 로직
    # ---------------------------------------------------------
    regions = ["서울", "경기", "인천", "부산", "대구", "광주", "대전"]
    sub_regions = {
        "서울": ["강남대로", "테헤란로", "도산대로", "여의도", "광화문", "명동", "홍대합정"],
        "경기": ["분당", "판교", "일산", "평촌"],
        "인천": ["부평", "구월", "송도"],
        "부산": ["서면", "해운대", "광복동"],
        "대구": ["동성로", "수성구"],
        "광주": ["상무지구", "충장로"],
        "대전": ["둔산", "은행동"]
    }
    asset_types = ["오피스", "소규모 상가", "중대형 상가"]
    quarters = ["2023 1Q", "2023 2Q", "2023 3Q", "2023 4Q", "2024 1Q", "2024 2Q"]
    
    data = []
    for r in regions:
        for sr in sub_regions[r]:
            for at in asset_types:
                for q in quarters:
                    base_rent = random.uniform(15000, 35000) if at == "오피스" else random.uniform(20000, 60000)
                    if r == "서울":
                        base_rent *= 1.5
                    vacancy = random.uniform(2.0, 15.0)
                    data.append({
                        "지역명(시/도)": r,
                        "세부 상권명": sr,
                        "자산 유형": at,
                        "기준 분기": q,
                        "㎡당 임대료": round(base_rent),
                        "공실률(%)": round(vacancy, 1)
                    })
    
    df = pd.DataFrame(data)
    df["평당 임대료"] = (df["㎡당 임대료"] * 3.3058).round().astype(int)
    return df

# ==========================================
# Tab 0: 마스터 대시보드

# ==========================================
with tab_master_dashboard:
    st.header("🌐 마스터 대시보드 (Executive Summary)")
    import plotly.express as px
    import numpy as np

    df_assets_md = fetch_data("SELECT * FROM Asset_Area")
    df_leases_md = fetch_data("SELECT * FROM Lease_Contracts WHERE status = 'ACTIVE'")

    if df_assets_md.empty:
        st.warning("등록된 자산 정보가 없습니다.")
    else:
        # [1단: 직관적 핵심 지표 (KPIs)]
        total_assets = df_assets_md["asset_name"].nunique()
        total_exclusive_area = df_assets_md["exclusive_area"].sum()
        total_bank_area = df_assets_md["bank_area"].sum() if "bank_area" in df_assets_md.columns else 0
        
        today_md = datetime.now()
        
        if not df_leases_md.empty:
            df_leases_md["start_date"] = pd.to_datetime(df_leases_md["start_date"])
            df_leases_md["end_date"] = pd.to_datetime(df_leases_md["end_date"])
            active_leases = df_leases_md[
                (df_leases_md["start_date"] <= today_md)
                & (df_leases_md["end_date"] >= today_md)
            ].copy()
        else:
            active_leases = pd.DataFrame()
            
        total_active_area = active_leases["contract_exclusive_area"].sum() if not active_leases.empty else 0
        
        # 1. 통합 관리 자산 및 임대율
        # 임대율 = (활성 임대 면적 + 은행 사용 면적) / 전체 전용 면적 * 100
        total_occupied = total_active_area + total_bank_area
        occupancy_rate = (total_occupied / total_exclusive_area * 100) if total_exclusive_area > 0 else 0
        
        # 2. 전체 임차사 및 활성 계약 수
        unique_companies = active_leases["company_name"].nunique() if not active_leases.empty else 0
        total_contracts = len(active_leases)
        
        # 3. 당월 총 청구 수익
        if not active_leases.empty:
            def get_krw_val(row, col):
                val = row[col] if pd.notnull(row[col]) else 0
                if row.get("currency", "KRW") == "USD":
                    return val * 1400
                return val

            active_leases["rent_krw"] = active_leases.apply(lambda r: get_krw_val(r, "monthly_rent"), axis=1)
            active_leases["maint_krw"] = active_leases.apply(lambda r: get_krw_val(r, "monthly_maintenance_fee"), axis=1)
            monthly_revenue = active_leases["rent_krw"].sum() + active_leases["maint_krw"].sum()
        else:
            monthly_revenue = 0

        # 4. 통합 공실 면적
        vacant_area = max(0, total_exclusive_area - total_occupied)

        st.markdown("### 📊 포트폴리오 핵심 지표 (Executive KPIs)")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric(
            "통합 관리 자산 및 임대율",
            f"{total_assets}개",
            f"임대율: {occupancy_rate:.1f}%",
        )
        c2.metric(
            "전체 임차사 및 활성 계약 수",
            f"{unique_companies}개 사",
            f"계약 {total_contracts}건",
        )
        mr_str = f"₩ {monthly_revenue/1000000:,.0f}백만" if monthly_revenue >= 1000000 else f"₩ {monthly_revenue:,.0f}"
        c3.metric(
            "당월 총 청구 수익",
            mr_str,
        )
        c4.metric(
            "통합 공실 면적",
            f"{vacant_area:,.1f} 평",
        )

        st.markdown("---")

        # [2단: 3대 핵심 시각화 차트]
        st.markdown("### 📈 통합 데이터 시각화 (Portfolio Analytics)")
        p1, p2 = st.columns(2)
        
        sc_palette = ["#005EB8", "#00A546", "#38BDF8", "#34D399", "#94A3B8"]

        with p1:
            with st.container(border=True):
                st.markdown("#### 포트폴리오 공간 점유 현황")
                pie_df = pd.DataFrame({
                    "Category": ["은행/지점 사용 면적", "일반 테넌트 임대 면적", "공실 면적"],
                    "Area": [total_bank_area, total_active_area, vacant_area]
                })
                pie_df = pie_df[pie_df["Area"] > 0]
                if not pie_df.empty:
                    fig_pie = px.pie(
                        pie_df, 
                        names="Category", 
                        values="Area", 
                        hole=0.4,
                        color_discrete_sequence=sc_palette
                    )
                    fig_pie.update_layout(
                        margin=dict(l=0, r=0, t=30, b=0),
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
                    )
                    fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                    st.plotly_chart(fig_pie, use_container_width=True)
                else:
                    st.info("데이터 없음")

        with p2:
            with st.container(border=True):
                st.markdown("#### 연도별 만기 도래 면적 (향후 5년)")
                if not active_leases.empty:
                    current_year = today_md.year
                    active_leases["end_year"] = active_leases["end_date"].dt.year
                    df_exp = active_leases[
                        (active_leases["end_year"] >= current_year) & 
                        (active_leases["end_year"] <= current_year + 5)
                    ]
                    if not df_exp.empty:
                        df_exp_grp = df_exp.groupby("end_year")["contract_area"].sum().reset_index()
                        df_exp_grp["end_year"] = df_exp_grp["end_year"].astype(str)
                        fig_exp = px.bar(
                            df_exp_grp,
                            x="end_year",
                            y="contract_area",
                            color_discrete_sequence=[sc_palette[0]]
                        )
                        fig_exp.update_layout(
                            xaxis_title="만기 연도",
                            yaxis_title="만기 도래 면적(평)",
                            margin=dict(l=0, r=0, t=30, b=0),
                            plot_bgcolor='rgba(0,0,0,0)',
                            paper_bgcolor='rgba(0,0,0,0)',
                            xaxis=dict(showgrid=False, zeroline=False),
                            yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False),
                            showlegend=False
                        )
                        fig_exp.update_traces(texttemplate='%{y:,.0f}', textposition='outside')
                        st.plotly_chart(fig_exp, use_container_width=True)
                    else:
                        st.info("향후 5년 내 만기 도래 계약 없음")
                else:
                    st.info("데이터 없음")
                    
        st.markdown("---")
        
        # [3단: 자산별 평단가 비교]
        with st.container(border=True):
            st.markdown("#### 자산별 평균 평단가 비교 (임대료 및 관리비)")
            if not active_leases.empty:
                valid_leases = active_leases[active_leases["contract_area"].astype(float) > 0].copy()
                valid_leases["rent_unit"] = valid_leases["rent_krw"] / valid_leases["contract_area"].astype(float)
                valid_leases["maint_unit"] = valid_leases["maint_krw"] / valid_leases["contract_area"].astype(float)
                df_unit = valid_leases.groupby("asset_name")[["rent_unit", "maint_unit"]].mean().reset_index()
                
                df_unit["total_unit"] = df_unit["rent_unit"] + df_unit["maint_unit"]
                df_unit = df_unit.sort_values(by="total_unit", ascending=False)
                
                df_melt = df_unit.melt(
                    id_vars=["asset_name"], 
                    value_vars=["rent_unit", "maint_unit"], 
                    var_name="Type", 
                    value_name="Unit_Price"
                )
                df_melt["Type"] = df_melt["Type"].replace({"rent_unit": "임대료", "maint_unit": "관리비"})
                
                fig_bar2 = px.bar(
                    df_melt,
                    x="asset_name",
                    y="Unit_Price",
                    color="Type",
                    barmode="group",
                    color_discrete_sequence=["#005EB8", "#34D399"]
                )
                fig_bar2.update_layout(
                    xaxis_title="자산명",
                    yaxis_title="평단가 (KRW)",
                    margin=dict(l=0, r=0, t=30, b=0),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=False, zeroline=False),
                    yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False),
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                fig_bar2.update_traces(texttemplate='₩ %{y:,.0f}', textposition='outside')
                st.plotly_chart(fig_bar2, use_container_width=True)
            else:
                st.info("데이터 없음")



# ==========================================
# Tab 0.5: 시장 동향 리서치
# ==========================================
with tab_market_research:
    st.header("📈 시장 동향 리서치 (한국부동산원 API 연동)")
    st.markdown("한국부동산원 상업용부동산 임대동향조사 오픈 API를 연동하여 주요 상권의 임대료 및 공실률 추이를 분석합니다. *(현재는 UI 데모용 가상 데이터를 표출 중입니다)*")
    
    market_df = fetch_market_research_data()
    
    f1, f2, f3 = st.columns(3)
    with f1:
        sel_regions = st.multiselect("📍 지역명(시/도)", options=market_df["지역명(시/도)"].unique(), default=["서울"])
    with f2:
        sel_types = st.multiselect("🏢 자산 유형", options=market_df["자산 유형"].unique(), default=["오피스"])
    with f3:
        latest_q = sorted(market_df["기준 분기"].unique())[-1]
        sel_quarters = st.multiselect("📅 기준 분기", options=market_df["기준 분기"].unique(), default=[latest_q])
        
    filtered_mdf = market_df.copy()
    if sel_regions:
        filtered_mdf = filtered_mdf[filtered_mdf["지역명(시/도)"].isin(sel_regions)]
    if sel_types:
        filtered_mdf = filtered_mdf[filtered_mdf["자산 유형"].isin(sel_types)]
    if sel_quarters:
        filtered_mdf = filtered_mdf[filtered_mdf["기준 분기"].isin(sel_quarters)]
        
    st.markdown("---")
    
    if not filtered_mdf.empty:
        agg_df = filtered_mdf.groupby("세부 상권명")[["평당 임대료", "공실률(%)"]].mean().reset_index()
        agg_df = agg_df.sort_values(by="평당 임대료", ascending=False)
        
        c1, c2 = st.columns(2)
        with c1:
            with st.container(border=True):
                st.markdown("#### 상권별 평균 평당 임대료")
                fig_rent = px.bar(agg_df, x="세부 상권명", y="평당 임대료")
                fig_rent.update_traces(marker_color="#005EB8", texttemplate='₩ %{y:,.0f}', textposition='outside')
                fig_rent.update_layout(
                    margin=dict(l=0, r=0, t=30, b=0),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=False, zeroline=False, title=""),
                    yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False)
                )
                st.plotly_chart(fig_rent, use_container_width=True)
                
        with c2:
            with st.container(border=True):
                st.markdown("#### 상권별 평균 공실률 (%)")
                fig_vac = px.bar(agg_df, x="세부 상권명", y="공실률(%)")
                fig_vac.update_traces(marker_color="#00A546", texttemplate='%{y:.1f}%', textposition='outside')
                fig_vac.update_layout(
                    margin=dict(l=0, r=0, t=30, b=0),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=False, zeroline=False, title=""),
                    yaxis=dict(showgrid=True, gridcolor="#F1F5F9", zeroline=False)
                )
                st.plotly_chart(fig_vac, use_container_width=True)
                
        st.markdown("---")
        st.markdown("#### 📋 상세 데이터 테이블")
        format_dict = {
            "㎡당 임대료": "₩ {:,.0f}",
            "평당 임대료": "₩ {:,.0f}",
            "공실률(%)": "{:.1f}%"
        }
        display_styled_table(filtered_mdf, freeze_cols=1, format_dict=format_dict)
    else:
        st.info("검색 조건에 일치하는 데이터가 없습니다.")

with tab_asset_view:

    st.header("자산별 면적 현황 조회")

    df_asset = fetch_data("SELECT * FROM Asset_Area")

    if not df_asset.empty:
        # Filter
        assets = df_asset["asset_name"].unique().tolist()
        selected_assets = st.multiselect(
            "🏢 자산명 필터 (미선택 시 전체 조회)", options=assets, default=[]
        )

        if selected_assets:
            df_asset = df_asset[df_asset["asset_name"].isin(selected_assets)]

        today_str = datetime.now().strftime("%Y-%m-%d")
        # 현재 활성화된 계약 면적 산출
        df_leases = fetch_data(
            f"SELECT asset_name, floor, contract_area FROM Lease_Contracts WHERE start_date <= '{today_str}' AND end_date >= '{today_str}' AND status = 'ACTIVE'"
        )

        if not df_leases.empty:
            leased_area_df = (
                df_leases.groupby(["asset_name", "floor"])["contract_area"]
                .sum()
                .reset_index()
            )
            leased_area_df.rename(
                columns={"contract_area": "leased_area"}, inplace=True
            )
            display_df = pd.merge(
                df_asset, leased_area_df, on=["asset_name", "floor"], how="left"
            )
            display_df["leased_area"] = display_df["leased_area"].fillna(0.0)
        else:
            display_df = df_asset.copy()
            display_df["leased_area"] = 0.0

        # Ensure bank_area is available
        if "bank_area" not in display_df.columns:
            display_df["bank_area"] = 0.0

        display_df["vacant_area"] = (
            display_df["exclusive_area"]
            - display_df["bank_area"].fillna(0.0)
            - display_df["leased_area"].fillna(0.0)
        )
        display_df["occupancy_rate (%)"] = (
            (
                display_df["leased_area"]
                / display_df["total_area"].replace(0, float("nan"))
                * 100
            )
            .fillna(0)
            .round(2)
        )
        display_df = sort_df_by_asset_and_floor(display_df, "asset_name", "floor")

        # 1. 자산별 토탈 대시보드
        dashboard_df = (
            display_df.groupby("asset_name")
            .agg(
                {
                    "total_area": "sum",
                    "common_area": "sum",
                    "exclusive_area": "sum",
                    "bank_area": "sum",
                    "leased_area": "sum",
                    "vacant_area": "sum",
                }
            )
            .reset_index()
        )

        dashboard_df["occupancy_rate (%)"] = (
            (
                dashboard_df["leased_area"]
                / dashboard_df["total_area"].replace(0, float("nan"))
                * 100
            )
            .fillna(0)
            .round(2)
        )
        dashboard_df = sort_df_by_asset_and_floor(dashboard_df, "asset_name", "floor")

        total_row_dash = pd.DataFrame(
            [
                {
                    "asset_name": "TOTAL",
                    "total_area": dashboard_df["total_area"].sum(),
                    "common_area": dashboard_df["common_area"].sum(),
                    "exclusive_area": dashboard_df["exclusive_area"].sum(),
                    "bank_area": dashboard_df["bank_area"].sum(),
                    "leased_area": dashboard_df["leased_area"].sum(),
                    "vacant_area": dashboard_df["vacant_area"].sum(),
                }
            ]
        )
        if total_row_dash["total_area"][0] > 0:
            total_row_dash["occupancy_rate (%)"] = round(
                (total_row_dash["leased_area"][0] / total_row_dash["total_area"][0])
                * 100,
                2,
            )
        else:
            total_row_dash["occupancy_rate (%)"] = 0.0

        dashboard_df = pd.concat([dashboard_df, total_row_dash], ignore_index=True)

        # Add Summary Row to details
        if not display_df.empty:
            summary = pd.DataFrame(
                [
                    {
                        "asset_name": "TOTAL",
                        "floor": "-",
                        "exclusive_area": display_df["exclusive_area"].sum(),
                        "common_area": display_df["common_area"].sum(),
                        "total_area": display_df["total_area"].sum(),
                        "bank_area": display_df["bank_area"].sum(),
                        "leased_area": display_df["leased_area"].sum(),
                        "vacant_area": display_df["vacant_area"].sum(),
                    }
                ]
            )
            if summary["total_area"][0] > 0:
                summary["occupancy_rate (%)"] = round(
                    (summary["leased_area"][0] / summary["total_area"][0]) * 100, 2
                )
            else:
                summary["occupancy_rate (%)"] = 0.0

            display_df = pd.concat([display_df, summary], ignore_index=True)

        st.write("")
        unit_option = st.radio(
            "🔄 표출 면적 단위 선택",
            ["평", "㎡", "sqft"],
            horizontal=True,
            key="tab1_unit_radio",
        )

        # Dashboard conversion
        dashboard_df_conv = dashboard_df.copy()
        dashboard_df_conv.rename(
            columns={
                "asset_name": "자산명",
                "total_area": "전체면적",
                "common_area": "공용면적",
                "bank_area": "은행 및 지점 사용 면적",
                "exclusive_area": "전용면적",
                "leased_area": "테넌트 면적",
                "vacant_area": "공실면적",
                "occupancy_rate (%)": "임대율 (%)",
            },
            inplace=True,
        )

        dashboard_order = [
            "자산명",
            "전체면적",
            "공용면적",
            "전용면적",
            "은행 및 지점 사용 면적",
            "테넌트 면적",
            "공실면적",
            "임대율 (%)",
        ]
        dashboard_df_conv = dashboard_df_conv[dashboard_order]

        area_cols = [
            "전체면적",
            "공용면적",
            "전용면적",
            "은행 및 지점 사용 면적",
            "테넌트 면적",
            "공실면적",
        ]

        if unit_option == "㎡":
            for col in area_cols:
                dashboard_df_conv[col] = (dashboard_df_conv[col] * 3.3058).round(2)
        elif unit_option == "sqft":
            for col in area_cols:
                dashboard_df_conv[col] = (dashboard_df_conv[col] * 35.58).round(2)

        csv_dash = dashboard_df_conv.to_csv(index=False).encode("utf-8-sig")
        file_name_dash = f"asset_total_dashboard_{unit_option}.csv"

        col_dash1, col_dash2, col_dash3 = st.columns([5, 3, 2])
        with col_dash1:
            st.markdown("### 📊 자산별 토탈 대시보드")
            st.download_button(
                "📊 토탈 CSV 다운로드",
                data=csv_dash,
                file_name=file_name_dash,
                mime="text/csv",
                use_container_width=True,
            )
            
        with col_dash2:
            to_email_dash = st.text_input("이메일", label_visibility="collapsed", placeholder="수신자 이메일 주소 입력", key="email_tab1_dash")
            
        with col_dash3:
            if st.button("🚀 메일 발송", key="btn_email_tab1_dash", use_container_width=True):
                if to_email_dash:
                    success, err = send_email_with_attachment(
                        to_email=to_email_dash,
                        subject="[PM/AM] 자산별 토탈 대시보드 리포트",
                        body="요청하신 자산별 토탈 대시보드 리포트 파일을 첨부하여 보내드립니다.",
                        file_bytes=csv_dash,
                        file_name=file_name_dash,
                        mime_type="text/csv"
                    )
                    if success:
                        st.toast("메일이 성공적으로 발송되었습니다!", icon="✅")
                    else:
                        st.error(f"메일 발송 실패: {err}")
                else:
                    st.warning("이메일 주소를 입력해주세요.")

        display_styled_table(
            center_styler(dashboard_df_conv).format(
                {c: "{:,.2f}" for c in area_cols + ["임대율 (%)"]}
            )
        )

        st.markdown("---")

        display_df_conv = display_df.copy()

        display_df_conv.rename(
            columns={
                "asset_name": "자산명",
                "floor": "해당층",
                "total_area": "전체면적",
                "common_area": "공용면적",
                "bank_area": "은행 및 지점 사용 면적",
                "exclusive_area": "전용면적",
                "leased_area": "테넌트 면적",
                "vacant_area": "공실면적",
                "occupancy_rate (%)": "임대율 (%)",
            },
            inplace=True,
        )

        # 열 배치 순서 변경: 전체면적 - 공용면적 - 전용면적 - 은행 및 지점 사용 면적 - 테넌트 면적 순
        desired_order = [
            "자산명",
            "해당층",
            "전체면적",
            "공용면적",
            "전용면적",
            "은행 및 지점 사용 면적",
            "테넌트 면적",
            "공실면적",
            "임대율 (%)",
        ]
        display_df_conv = display_df_conv[desired_order]

        area_cols = [
            "전체면적",
            "공용면적",
            "전용면적",
            "은행 및 지점 사용 면적",
            "테넌트 면적",
            "공실면적",
        ]

        if unit_option == "㎡":
            for col in area_cols:
                display_df_conv[col] = (display_df_conv[col] * 3.3058).round(2)
        elif unit_option == "sqft":
            for col in area_cols:
                display_df_conv[col] = (display_df_conv[col] * 35.58).round(2)

        csv = display_df_conv.to_csv(index=False).encode("utf-8-sig")
        file_name_1 = f"asset_area_status_{unit_option}.csv"

        col_a1, col_a2, col_a3 = st.columns([5, 3, 2])
        with col_a1:
            st.markdown("### 🏢 자산별 층별 상세 현황")
            st.download_button(
                "📊 현황 CSV 다운로드",
                data=csv,
                file_name=file_name_1,
                mime="text/csv",
                use_container_width=True,
            )
            
        with col_a2:
            to_email_1 = st.text_input("이메일", label_visibility="collapsed", placeholder="수신자 이메일 주소 입력", key="email_tab1")
            
        with col_a3:
            if st.button("🚀 메일 발송", key="btn_email_tab1", use_container_width=True):
                if to_email_1:
                    success, err = send_email_with_attachment(
                        to_email=to_email_1,
                        subject="[PM/AM] 자산별 면적 현황 리포트",
                        body="요청하신 자산별 면적 현황 리포트 파일을 첨부하여 보내드립니다.",
                        file_bytes=csv,
                        file_name=file_name_1,
                        mime_type="text/csv"
                    )
                    if success:
                        st.toast("메일이 성공적으로 발송되었습니다!", icon="✅")
                    else:
                        st.error(f"메일 발송 실패: {err}")
                else:
                    st.warning("이메일 주소를 입력해주세요.")

        display_styled_table(
            center_styler(display_df_conv).format(
                {c: "{:,.2f}" for c in area_cols + ["임대율 (%)"]}
            )
        )
    else:
        st.info(
            "등록된 자산 정보가 없습니다. '자산정보 업데이트' 탭에서 데이터를 입력해 주세요."
        )


# ==========================================
# Tab 1.5: 스태킹 플랜
# ==========================================
with tab_stacking_plan:
    st.header("자산별 스태킹 플랜 (Visual Stacking Plan)")

    # Needs assets list
    df_asset_sp = fetch_data("SELECT DISTINCT asset_name FROM Asset_Area")
    if not df_asset_sp.empty:
        assets_sp = df_asset_sp["asset_name"].tolist()
    # --- Visual Stacking Plan ---
    st.markdown("---")
    st.subheader("🏢 자산별 스태킹 플랜 (Visual Stacking Plan)")

    # Select single asset for stacking plan
    sp_asset = st.selectbox("스태킹 플랜을 조회할 자산을 선택하세요", options=assets)

    if sp_asset:
        unit_sp = st.radio(
            "🔄 표출 면적 단위 선택",
            ["평", "㎡", "sqft"],
            horizontal=True,
            key="sp_unit_radio",
        )
        mult = 1.0
        if unit_sp == "㎡":
            mult = 3.3058
        elif unit_sp == "sqft":
            mult = 35.58

        # Get floors for this asset
        df_floors = fetch_data(
            f"SELECT floor, exclusive_area, bank_area FROM Asset_Area WHERE asset_name = '{sp_asset}'"
        )
        df_leases_sp = fetch_data(
            f"SELECT floor, company_name, contract_area FROM Lease_Contracts WHERE asset_name = '{sp_asset}' AND status = 'ACTIVE' AND start_date <= '{today_str}' AND end_date >= '{today_str}'"
        )

        # Sort floors dynamically
        def floor_sort_key(f):
            f = str(f).upper()
            if f.startswith("B"):
                try:
                    return -int("".join(filter(str.isdigit, f)))
                except:
                    return -99
            else:
                try:
                    return int("".join(filter(str.isdigit, f)))
                except:
                    return 0

        df_floors["sort_key"] = df_floors["floor"].apply(floor_sort_key)
        df_floors = df_floors.sort_values("sort_key", ascending=False)

        st.write(f"**{sp_asset}** 층별 입주 현황")

        for _, floor_row in df_floors.iterrows():
            floor_name = floor_row["floor"]
            exclusive = floor_row["exclusive_area"] * mult
            bank_area = floor_row.get("bank_area", 0.0)
            if pd.isna(bank_area):
                bank_area = 0.0
            bank_area *= mult

            # Find tenants on this floor
            floor_leases = df_leases_sp[df_leases_sp["floor"] == floor_name].copy()
            floor_leases["contract_area"] = floor_leases["contract_area"] * mult

            blocks_html = ""

            # Bank block
            if bank_area > 0:
                flex_val = max(0.1, bank_area)
                blocks_html += f"<div title='은행/지점&#10;면적: {bank_area:.1f} {unit_sp}' style='flex: {flex_val}; background-color: #005EB8; color: white; padding: 10px; margin: 2px; border-radius: 4px; text-align: center; min-width: 50px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'><b>은행/지점</b><br>{bank_area:.1f}</div>"

            leased_total = 0.0
            for _, l_row in floor_leases.iterrows():
                comp = l_row["company_name"]
                area = (
                    float(l_row["contract_area"])
                    if not pd.isna(l_row["contract_area"])
                    else 0.0
                )
                leased_total += area

                if "은행" in comp or "SC" in comp.upper() or "BANK" in comp.upper():
                    bg_color = "#005EB8"  # SC Blue
                    text_color = "white"
                else:
                    bg_color = "#00A546"  # SC Green
                    text_color = "white"

                if area > 0:
                    flex_val = max(0.1, area)
                    blocks_html += f"<div title='임차사: {comp}&#10;면적: {area:.1f} {unit_sp}' style='flex: {flex_val}; background-color: {bg_color}; color: {text_color}; padding: 10px; margin: 2px; border-radius: 4px; text-align: center; min-width: 50px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'><b>{comp}</b><br>{area:.1f}</div>"

            vacant = float(exclusive) - float(bank_area) - leased_total
            if pd.isna(vacant):
                vacant = 0.0

            if vacant > 0.1:
                flex_val = max(0.1, vacant)
                blocks_html += f"<div title='공실&#10;면적: {vacant:.1f} {unit_sp}' style='flex: {flex_val}; background-color: #9ca3af; color: white; padding: 10px; margin: 2px; border-radius: 4px; text-align: center; min-width: 50px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'><b>공실</b><br>{vacant:.1f}</div>"

            if blocks_html == "":
                # 면적이 0이거나 데이터가 없는 층 (RF 등)
                blocks_html = f"<div title='면적 0 / 데이터 없음' style='flex: 1; background-color: #e5e7eb; color: #6b7280; padding: 10px; margin: 2px; border-radius: 4px; text-align: center; min-width: 50px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'><b>0</b></div>"

            row_html = f"""
            <div style="display: flex; align-items: stretch; margin-bottom: 5px; border: 1px solid #e5e7eb; padding: 5px; background-color: #f9fafb; border-radius: 4px;">
                <div style="width: 80px; display: flex; align-items: center; justify-content: center; font-weight: bold; background-color: #f3f4f6; margin-right: 10px; border-radius: 4px;">
                    {floor_name}
                </div>
                <div style="display: flex; flex: 1;">
                    {blocks_html}
                </div>
            </div>
            """
            st.markdown(row_html, unsafe_allow_html=True)

    else:
        st.info("등록된 자산 정보가 없습니다.")

# ==========================================
# Tab 2: 자산별 임대정보 관리
# ==========================================
with tab_lease_info:
    st.header("자산별 임대정보 관리")
    st.info(
        "※ 현재 'ACTIVE' 상태인 활성 계약들만 노출됩니다. 퇴점(TERMINATED) 처리된 계약 및 과거 계약(갱신 완료로 인한 이전 기록)은 중복 표출 방지를 위해 숨김 처리됩니다."
    )

    # Only fetch ACTIVE contracts for current leasing info
    df_contracts = fetch_data(
        "SELECT * FROM Lease_Contracts WHERE status = 'ACTIVE' OR status IS NULL"
    )

    if not df_contracts.empty:
        # Filters
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            assets = df_contracts["asset_name"].unique().tolist()
            sel_assets = st.multiselect(
                "🏢 자산명 필터", options=assets, default=[], key="tab2_asset_filter"
            )
        with col_f2:
            companies = df_contracts["company_name"].unique().tolist()
            sel_companies = st.multiselect(
                "🏢 업체명 필터",
                options=companies,
                default=[],
                key="tab2_company_filter",
            )

        if sel_assets:
            df_contracts = df_contracts[df_contracts["asset_name"].isin(sel_assets)]
        if sel_companies:
            df_contracts = df_contracts[
                df_contracts["company_name"].isin(sel_companies)
            ]

        def calc_effective_rent(row):
            try:
                start = pd.to_datetime(row["start_date"])
                end = pd.to_datetime(row["end_date"])

                total_days = (end - start).days + 1
                total_months = total_days / 365 * 12

                if total_months <= 0:
                    return 0
                rent = float(row["monthly_rent"])
                rf_months = float(row["total_rent_free_months"])
                return ((rent * total_months) - (rent * rf_months)) / total_months
            except:
                return 0

        df_contracts["effective_rent"] = df_contracts.apply(calc_effective_rent, axis=1)

        df_contracts["deposit_per_pyeong"] = (
            df_contracts["deposit"]
            / df_contracts["contract_area"].astype(float).replace(0, float("nan"))
        ).fillna(0)
        df_contracts["rent_per_pyeong"] = (
            df_contracts["monthly_rent"]
            / df_contracts["contract_area"].astype(float).replace(0, float("nan"))
        ).fillna(0)
        df_contracts["maintenance_per_pyeong"] = (
            df_contracts["monthly_maintenance_fee"]
            / df_contracts["contract_area"].astype(float).replace(0, float("nan"))
        ).fillna(0)

        display_cols = [
            "company_name",
            "asset_name",
            "floor",
            "start_date",
            "end_date",
            "contract_area",
            "contract_exclusive_area",
            "currency",
            "deposit",
            "deposit_per_pyeong",
            "monthly_rent",
            "rent_per_pyeong",
            "monthly_maintenance_fee",
            "maintenance_per_pyeong",
            "total_rent_free_months",
            "effective_rent",
            "remarks",
            "contract_id",
        ]

        df_display = df_contracts[display_cols].copy()

        # Ensure currency has a default
        df_display["currency"] = df_display["currency"].fillna("KRW")
        df_display = sort_df_by_asset_and_floor(df_display, "asset_name", "floor")

        rename_dict = {
            "contract_id": "계약ID",
            "asset_name": "자산명",
            "floor": "층",
            "company_name": "업체명",
            "start_date": "계약시작일",
            "end_date": "계약종료일",
            "contract_area": "계약면적(평)",
            "contract_exclusive_area": "전용면적(평)",
            "currency": "통화",
            "deposit": "보증금",
            "deposit_per_pyeong": "평당 보증금",
            "monthly_rent": "임대료",
            "rent_per_pyeong": "평당 임대료",
            "monthly_maintenance_fee": "관리비",
            "maintenance_per_pyeong": "평당 관리비",
            "total_rent_free_months": "렌트프리(개월)",
            "effective_rent": "실질 임대료",
            "remarks": "비고",
        }
        df_display.rename(columns=rename_dict, inplace=True)

        df_krw = df_display[df_display["통화"] == "KRW"].copy()
        df_usd = df_display[df_display["통화"] == "USD"].copy()

        csv2 = df_display.to_csv(index=False).encode("utf-8-sig")
        file_name_2 = "lease_contracts.csv"

        col_sum1, col_sum2, col_sum3 = st.columns([5, 3, 2])
        with col_sum1:
            st.markdown("### 📊 자산 통합 Summary")
            st.download_button(
                "📝 전체 통합 CSV 다운로드",
                data=csv2,
                file_name=file_name_2,
                mime="text/csv",
                use_container_width=True,
            )
            
        with col_sum2:
            to_email_2 = st.text_input("이메일", label_visibility="collapsed", placeholder="수신자 이메일 주소 입력", key="email_tab2")
            
        with col_sum3:
            if st.button("🚀 메일 발송", key="btn_email_tab2", use_container_width=True):
                if to_email_2:
                    success, err = send_email_with_attachment(
                        to_email=to_email_2,
                        subject="[PM/AM] 통합 임대정보 리포트",
                        body="요청하신 통합 임대정보 리포트 파일을 첨부하여 보내드립니다.",
                        file_bytes=csv2,
                        file_name=file_name_2,
                        mime_type="text/csv"
                    )
                    if success:
                        st.toast("메일이 성공적으로 발송되었습니다!", icon="✅")
                    else:
                        st.error(f"메일 발 발송 실패: {err}")
                else:
                    st.warning("이메일 주소를 입력해주세요.")
        sum_krw_dep = df_krw["보증금"].sum() if not df_krw.empty else 0
        sum_krw_rent = df_krw["임대료"].sum() if not df_krw.empty else 0
        sum_krw_maint = df_krw["관리비"].sum() if not df_krw.empty else 0

        sum_usd_dep = df_usd["보증금"].sum() if not df_usd.empty else 0
        sum_usd_rent = df_usd["임대료"].sum() if not df_usd.empty else 0
        sum_usd_maint = df_usd["관리비"].sum() if not df_usd.empty else 0

        st.markdown(
            f"""
        <div style="display: flex; gap: 1rem; margin-bottom: 1rem;">
            <div style="background-color: white; padding: 1rem; border-radius: 8px; box-shadow: 0 1px 2px rgba(0,0,0,0.05); flex: 1; text-align: center;">
                <p style="color: #64748b; font-size: 0.8rem; margin: 0; font-weight: 600;">총 보증금</p>
                <p style="color: #1e293b; font-size: 1.6rem; margin: 0; font-weight: 700;">₩ {sum_krw_dep:,.0f} <span style="font-size: 1rem; color: #94a3b8;">/ USD {sum_usd_dep:,.2f}</span></p>
            </div>
            <div style="background-color: white; padding: 1rem; border-radius: 8px; box-shadow: 0 1px 2px rgba(0,0,0,0.05); flex: 1; text-align: center;">
                <p style="color: #64748b; font-size: 0.8rem; margin: 0; font-weight: 600;">총 임대료</p>
                <p style="color: #1e293b; font-size: 1.6rem; margin: 0; font-weight: 700;">₩ {sum_krw_rent:,.0f} <span style="font-size: 1rem; color: #94a3b8;">/ USD {sum_usd_rent:,.2f}</span></p>
            </div>
            <div style="background-color: white; padding: 1rem; border-radius: 8px; box-shadow: 0 1px 2px rgba(0,0,0,0.05); flex: 1; text-align: center;">
                <p style="color: #64748b; font-size: 0.8rem; margin: 0; font-weight: 600;">총 관리비</p>
                <p style="color: #1e293b; font-size: 1.6rem; margin: 0; font-weight: 700;">₩ {sum_krw_maint:,.0f} <span style="font-size: 1rem; color: #94a3b8;">/ USD {sum_usd_maint:,.2f}</span></p>
            </div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        st.markdown("---")

        currency_cols_kor = [
            "보증금",
            "평당 보증금",
            "임대료",
            "평당 임대료",
            "관리비",
            "평당 관리비",
            "실질 임대료",
        ]

        if not df_krw.empty:
            st.markdown(
                "<p style='font-size: 0.9rem; font-weight: bold; margin-bottom: 0.5rem;'>계약 내역</p>",
                unsafe_allow_html=True,
            )
            display_styled_table(
                center_styler(df_krw).format(
                    {
                        **{c: "₩ {:,.0f}" for c in currency_cols_kor},
                        "계약면적(평)": "{:,.2f}",
                        "전용면적(평)": "{:,.2f}",
                    }
                )
            )

        if not df_usd.empty:
            st.markdown(
                "<p style='font-size: 0.9rem; font-weight: bold; margin-bottom: 0.5rem;'>USD 계약 내역</p>",
                unsafe_allow_html=True,
            )
            display_styled_table(
                center_styler(df_usd).format(
                    {
                        **{c: "USD {:,.2f}" for c in currency_cols_kor},
                        "계약면적(평)": "{:,.2f}",
                        "전용면적(평)": "{:,.2f}",
                    }
                )
            )

    else:
        st.info("등록된 활성 계약 정보가 없습니다.")

# ==========================================
# Tab 3: 렌트롤 관리
# ==========================================

with tab_rent_roll:
    st.header("렌트롤 (Rent Roll) 관리 및 수동 조정")
    st.markdown(
        "자동 산출된 금액 외에 예외적으로 조정이 필요한 달이 있다면 표의 금액을 직접 클릭하여 **수정 후 [저장]** 할 수 있습니다. 갱신/퇴점된 계약 이력도 해당 날짜에 맞춰 정상적으로 병합 표출됩니다."
    )

    # 렌트롤은 모든 상태의 계약 내역을 기반으로 산출
    df_c = fetch_data("SELECT * FROM Lease_Contracts")

    if not df_c.empty:
        col_f1, col_f2, col_y1 = st.columns(3)
        with col_f1:
            assets = df_c["asset_name"].unique().tolist()
            sel_assets = st.multiselect(
                "🏢 자산명 필터", options=assets, default=[], key="tab3_asset_filter"
            )
        with col_f2:
            companies = df_c["company_name"].unique().tolist()
            sel_companies = st.multiselect(
                "🏢 업체명 필터",
                options=companies,
                default=[],
                key="tab3_company_filter",
            )
        with col_y1:
            start_year = 2026
            selected_year = st.selectbox(
                "📅 조회 연도 선택",
                range(start_year, max(datetime.now().year, start_year) + 10),
                index=0,
            )

        if sel_assets:
            df_c = df_c[df_c["asset_name"].isin(sel_assets)]
        if sel_companies:
            df_c = df_c[df_c["company_name"].isin(sel_companies)]

        # Load Overrides
        df_overrides = fetch_data(
            f"SELECT * FROM RentRoll_Overrides WHERE year = {selected_year}"
        )
        overrides_dict = {}
        for _, ov in df_overrides.iterrows():
            overrides_dict[(ov["contract_id"], ov["floor"], ov["month"])] = (
                ov["over_rent"],
                ov["over_maint"],
            )

        records = []
        for _, row in df_c.iterrows():
            try:
                start = pd.to_datetime(row["start_date"])
                end = pd.to_datetime(row["end_date"])

                year_start = datetime(selected_year, 1, 1)
                year_end = datetime(selected_year, 12, 31)
                if start > year_end or end < year_start:
                    continue

                rf_details = (
                    json.loads(row["rent_free_details"])
                    if row["rent_free_details"]
                    else []
                )

                status_str = f"[{row['status']}] " if row["status"] != "ACTIVE" else ""

                # Parse floor details
                fd_raw = (
                    row["floor_details"]
                    if "floor_details" in row and pd.notnull(row["floor_details"])
                    else None
                )
                if fd_raw:
                    try:
                        floors_info = json.loads(fd_raw)
                    except:
                        floors_info = {row["floor"]: {"ratio": 1.0}}
                else:
                    floors_info = {row["floor"]: {"ratio": 1.0}}

                # Create a record for each floor
                floor_records = {}
                for fl in floors_info.keys():
                    floor_records[fl] = {
                        "Contract_ID": row["contract_id"],
                        "자산명": row["asset_name"],
                        "층": fl,
                        "업체명": status_str + row["company_name"],
                        "통화": (
                            row["currency"]
                            if "currency" in row and pd.notnull(row["currency"])
                            else "KRW"
                        ),
                    }

                start_month = 6 if selected_year == 2026 else 1
                for month in range(start_month, 13):
                    month_str = f"{selected_year}-{month:02d}"
                    _, last_day = calendar.monthrange(selected_year, month)
                    curr_month_start = datetime(selected_year, month, 1)
                    curr_month_end = datetime(selected_year, month, last_day)

                    # 1. Escalation Logic
                    escalated_rent, escalated_maint = calculate_escalated_amount(
                        row, curr_month_start
                    )

                    # 2. Floor to 10 won (already handled in calculate_escalated_amount for KRW)
                    rent_rounded = escalated_rent
                    maint_rounded = escalated_maint

                    overlap_start = max(start, curr_month_start)
                    overlap_end = min(end, curr_month_end)

                    is_rf = month_str in rf_details
                    actual_rent_total = 0 if is_rf else rent_rounded
                    actual_maint_total = maint_rounded

                    if overlap_start <= overlap_end:
                        days_in_overlap = (overlap_end - overlap_start).days + 1
                        if days_in_overlap < last_day:
                            rent_to_charge_total = (
                                actual_rent_total / last_day
                            ) * days_in_overlap
                            maint_to_charge_total = (
                                actual_maint_total / last_day
                            ) * days_in_overlap
                        else:
                            rent_to_charge_total = actual_rent_total
                            maint_to_charge_total = actual_maint_total
                    else:
                        rent_to_charge_total = 0
                        maint_to_charge_total = 0

                    # 3. Distribute by floor
                    for fl, info in floors_info.items():
                        if (row["contract_id"], fl, month) in overrides_dict:
                            o_rent, o_maint = overrides_dict[(row["contract_id"], fl, month)]
                            floor_rent = o_rent
                            floor_maint = o_maint
                        else:
                            ratio = info.get("ratio", 1.0)
                            floor_rent = rent_to_charge_total * ratio
                            floor_maint = maint_to_charge_total * ratio

                        floor_records[fl][f"{month}월 임대료"] = round(floor_rent)
                        floor_records[fl][f"{month}월 관리비"] = round(floor_maint)

                for fl, rec in floor_records.items():
                    records.append(rec)

            except Exception as e:
                st.error(
                    f"데이터 처리 중 오류 발생 (Contract ID: {row['contract_id']}): {e}"
                )

        if records:
            df_rr = pd.DataFrame(records)
            df_rr = sort_df_by_asset_and_floor(df_rr, "자산명", "층")

            df_rr_krw = df_rr[df_rr["통화"] == "KRW"].copy()
            df_rr_usd = df_rr[df_rr["통화"] == "USD"].copy()

            csv_rr = df_rr.to_csv(index=False).encode("utf-8-sig")
            file_name_3 = f"rent_roll_{selected_year}_details.csv"

            col_r1, col_r2, col_r3 = st.columns([5, 3, 2])
            with col_r1:
                st.markdown(f"### {selected_year}년 렌트롤 상세 내역")
                st.download_button(
                    "📥 통합 렌트롤 CSV 다운로드",
                    data=csv_rr,
                    file_name=file_name_3,
                    mime="text/csv",
                    use_container_width=True,
                )
                
            with col_r2:
                to_email_3 = st.text_input("이메일", label_visibility="collapsed", placeholder="수신자 이메일 주소 입력", key="email_tab3")
                
            with col_r3:
                if st.button("🚀 메일 발송", key="btn_email_tab3", use_container_width=True):
                    if to_email_3:
                        success, err = send_email_with_attachment(
                            to_email=to_email_3,
                            subject=f"[PM/AM] {selected_year}년 렌트롤 리포트",
                            body=f"요청하신 {selected_year}년 렌트롤 상세 내역을 첨부하여 보내드립니다.",
                            file_bytes=csv_rr,
                            file_name=file_name_3,
                            mime_type="text/csv"
                        )
                        if success:
                            st.toast("메일이 성공적으로 발송되었습니다!", icon="✅")
                        else:
                            st.error(f"메일 발송 실패: {err}")
                    else:
                        st.warning("이메일 주소를 입력해주세요.")

            view_mode = st.radio(
                "보기 모드 선택",
                ["👁️ 조회 모드 (완벽한 디자인 적용)", "📝 예외 숫자 편집 모드"],
                horizontal=True,
            )

            if view_mode == "👁️ 조회 모드 (완벽한 디자인 적용)":
                rr_css = """
                .custom-st-table.{uid} tr:nth-child(even) td:nth-child(odd):nth-child(n+5) { background-color: #e0f2fe !important; }
                .custom-st-table.{uid} tr:nth-child(odd) td:nth-child(odd):nth-child(n+5) { background-color: #f0f9ff !important; }
                .custom-st-table.{uid} th:nth-child(odd):nth-child(n+5) { background-color: #e0f2fe !important; }

                .custom-st-table.{uid} tr:nth-child(even) td:nth-child(even):nth-child(n+6) { background-color: #fef3c7 !important; }
                .custom-st-table.{uid} tr:nth-child(odd) td:nth-child(even):nth-child(n+6) { background-color: #fffbeb !important; }
                .custom-st-table.{uid} th:nth-child(even):nth-child(n+6) { background-color: #fef3c7 !important; }

                .custom-st-table.{uid} td:nth-child(even):nth-child(n+6), 
                .custom-st-table.{uid} th:nth-child(even):nth-child(n+6) {
                    border-right: 2px solid #64748b !important;
                }
                """
                if not df_rr_krw.empty:
                    st.markdown("#### 🇰🇷 KRW 렌트롤")
                    format_dict_krw = {}
                    start_m = 6 if selected_year == 2026 else 1

                    def fmt_krw(x):
                        if pd.isna(x):
                            return ""
                        try:
                            val = float(x)
                            if abs(val) >= 1_000_000:
                                return f"₩ {val/1_000_000:,.1f}백만"
                            return f"₩ {val:,.0f}"
                        except:
                            return str(x)

                    for m in range(start_m, 13):
                        format_dict_krw[f"{m}월 임대료"] = fmt_krw
                        format_dict_krw[f"{m}월 관리비"] = fmt_krw
                    display_styled_table(
                        df_rr_krw.drop(columns=["Contract_ID"]),
                        freeze_cols=4,
                        format_dict=format_dict_krw,
                        custom_css=rr_css,
                    )

                if not df_rr_usd.empty:
                    st.markdown("#### 🇺🇸 USD 렌트롤")
                    format_dict_usd = {}
                    start_m = 6 if selected_year == 2026 else 1

                    def fmt_usd(x):
                        if pd.isna(x):
                            return ""
                        try:
                            val = float(x)
                            if abs(val) >= 1_000_000:
                                return f"USD {val/1_000_000:,.2f}백만"
                            return f"USD {val:,.2f}"
                        except:
                            return str(x)

                    for m in range(start_m, 13):
                        format_dict_usd[f"{m}월 임대료"] = fmt_usd
                        format_dict_usd[f"{m}월 관리비"] = fmt_usd
                    display_styled_table(
                        df_rr_usd.drop(columns=["Contract_ID"]),
                        freeze_cols=4,
                        format_dict=format_dict_usd,
                        custom_css=rr_css,
                    )
            else:
                st.info(
                    "※ 숫자를 더블클릭하여 수정하신 후, 반드시 아래의 [저장] 버튼을 눌러주세요. 수정 모드에서는 스트림릿 기본 디자인만 지원됩니다."
                )
                disabled_cols = ["Contract_ID", "자산명", "층", "업체명", "통화"]
                col_config_base = {
                    col: st.column_config.Column(disabled=True) for col in disabled_cols
                }

                edited_krw = None
                edited_usd = None

                if not df_rr_krw.empty:
                    st.markdown("#### 🇰🇷 KRW 렌트롤")
                    col_config_krw = col_config_base.copy()
                    start_m = 6 if selected_year == 2026 else 1
                    for month in range(start_m, 13):
                        col_r = f"{month}월 임대료"
                        col_m = f"{month}월 관리비"
                        df_rr_krw[col_r] = df_rr_krw[col_r].apply(
                            lambda x: f"{int(float(x)):,}"
                        )
                        df_rr_krw[col_m] = df_rr_krw[col_m].apply(
                            lambda x: f"{int(float(x)):,}"
                        )
                        col_config_krw[col_r] = st.column_config.TextColumn(
                            f"₩ {col_r}"
                        )
                        col_config_krw[col_m] = st.column_config.TextColumn(
                            f"₩ {col_m}"
                        )

                    index_cols = ["Contract_ID", "자산명", "층", "업체명", "통화"]
                    edited_krw = st.data_editor(
                        df_rr_krw,
                        use_container_width=True,
                        column_config=col_config_krw,
                        hide_index=True,
                        key=f"rr_editor_krw_{selected_year}",
                    )

                if not df_rr_usd.empty:
                    st.markdown("#### 🇺🇸 USD 렌트롤")
                    col_config_usd = col_config_base.copy()
                    start_m = 6 if selected_year == 2026 else 1
                    for month in range(start_m, 13):
                        col_r = f"{month}월 임대료"
                        col_m = f"{month}월 관리비"
                        df_rr_usd[col_r] = df_rr_usd[col_r].apply(
                            lambda x: f"{float(x):,.2f}"
                        )
                        df_rr_usd[col_m] = df_rr_usd[col_m].apply(
                            lambda x: f"{float(x):,.2f}"
                        )
                        col_config_usd[col_r] = st.column_config.TextColumn(
                            f"USD {col_r}"
                        )
                        col_config_usd[col_m] = st.column_config.TextColumn(
                            f"USD {col_m}"
                        )

                    index_cols = ["Contract_ID", "자산명", "층", "업체명", "통화"]
                    edited_usd = st.data_editor(
                        df_rr_usd,
                        use_container_width=True,
                        column_config=col_config_usd,
                        hide_index=True,
                        key=f"rr_editor_usd_{selected_year}",
                    )

                if st.button("💾 렌트롤 예외 수정사항 DB에 저장", type="primary"):
                    db_conn = engine.raw_connection()
                    c = db_conn.cursor()
                    changes_made = 0

                    def process_edited(df_edited, df_orig):
                        cnt = 0
                        if df_edited is None or df_edited.empty:
                            return cnt
                        for idx in df_edited.index:
                            start_m = 6 if selected_year == 2026 else 1
                            for month in range(start_m, 13):
                                rent_col = f"{month}월 임대료"
                                maint_col = f"{month}월 관리비"

                                new_rent = df_edited.loc[idx, rent_col]
                                new_maint = df_edited.loc[idx, maint_col]
                                old_rent = df_orig.loc[idx, rent_col]
                                old_maint = df_orig.loc[idx, maint_col]

                                if str(new_rent) != str(old_rent) or str(
                                    new_maint
                                ) != str(old_maint):
                                    cid = int(df_edited.loc[idx, "Contract_ID"])
                                    fl = str(df_edited.loc[idx, "층"])
                                    val_rent = float(str(new_rent).replace(",", "").replace("₩", "").replace("USD", "").replace(" ", ""))
                                    val_maint = float(str(new_maint).replace(",", "").replace("₩", "").replace("USD", "").replace(" ", ""))
                                    c.execute(
                                        """
                                        INSERT INTO RentRoll_Overrides (contract_id, floor, year, month, over_rent, over_maint)
                                        VALUES (%s, %s, %s, %s, %s, %s)
                                        ON CONFLICT(contract_id, floor, year, month)
                                        DO UPDATE SET over_rent=excluded.over_rent, over_maint=excluded.over_maint
                                    """,
                                        (
                                            cid,
                                            fl,
                                            selected_year,
                                            month,
                                            val_rent,
                                            val_maint,
                                        ),
                                    )
                                    cnt += 1
                        return cnt

                    changes_made += process_edited(edited_krw, df_rr_krw)
                    changes_made += process_edited(edited_usd, df_rr_usd)

                    if changes_made > 0:
                        db_conn.commit()
                        db_conn.close()
                        fetch_data.clear()
                        st.success(
                            f"✅ {changes_made}건의 렌트롤 예외 수정사항이 데이터베이스에 안전하게 저장되었습니다."
                        )
                    else:
                        st.info("수정된 항목이 없습니다.")

    else:
        st.info("해당 연도에 포함된 계약 정보가 없습니다.")

# ==========================================
# Tab 4: 자산정보 업데이트
# ==========================================
with tab_asset_update:
    st.header("자산정보 등록 및 업데이트")

    st.markdown("### 1. CSV 일괄 업로드")

    template_df = pd.DataFrame(
        columns=[
            "자산명",
            "해당층",
            "전체면적",
            "공용면적",
            "전용면적",
            "은행 및 지점 사용 면적",
        ]
    )
    template_csv = template_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "📝 빈 양식 다운로드 (CSV)",
        data=template_csv,
        file_name="asset_upload_template.csv",
        mime="text/csv",
    )

    st.info(
        "※ 면적 단위는 **'평'**을 기준으로 입력하여 업로드해주세요. 기존에 동일한 자산/층이 있다면 덮어쓰기(업데이트) 됩니다."
    )
    uploaded_file = st.file_uploader(
        "작성된 자산정보 CSV 파일 선택", type=["csv", "xlsx", "xls"]
    )
    if uploaded_file:
        try:
            if uploaded_file.name.endswith(".csv"):
                df_upload = pd.read_csv(uploaded_file)
            else:
                df_upload = pd.read_excel(uploaded_file)

            st.markdown("#### 업로드된 데이터 미리보기")
            num_cols = [
                "전체면적",
                "공용면적",
                "전용면적",
                "은행 및 지점 사용 면적",
                "exclusive_area",
                "common_area",
                "total_area",
                "bank_area",
            ]

            for c in num_cols:
                if c in df_upload.columns:
                    if pd.api.types.is_string_dtype(df_upload[c]):
                        df_upload[c] = df_upload[c].str.replace(",", "")
                    df_upload[c] = pd.to_numeric(df_upload[c], errors="coerce").fillna(
                        0.0
                    )

            display_styled_table(
                center_styler(df_upload).format(
                    {c: "{:,.2f}" for c in num_cols if c in df_upload.columns}
                )
            )

            if st.button("✅ 데이터베이스에 최종 반영하기", type="primary"):
                df_to_process = df_upload.rename(
                    columns={
                        "자산명": "asset_name",
                        "해당층": "floor",
                        "전체면적": "total_area",
                        "공용면적": "common_area",
                        "전용면적": "exclusive_area",
                        "은행 및 지점 사용 면적": "bank_area",
                    }
                )

                db_conn = engine.raw_connection()
                c = db_conn.cursor()
                for _, row in df_to_process.iterrows():
                    if all(
                        col in row
                        for col in [
                            "asset_name",
                            "floor",
                            "exclusive_area",
                            "common_area",
                            "total_area",
                        ]
                    ):
                        bank_area = float(
                            row.get("bank_area", 0.0)
                            if not pd.isna(row.get("bank_area", 0.0))
                            else 0.0
                        )
                        c.execute(
                            """
                            INSERT INTO Asset_Area (asset_name, floor, exclusive_area, common_area, total_area, bank_area)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            ON CONFLICT(asset_name, floor) 
                            DO UPDATE SET 
                                exclusive_area=excluded.exclusive_area, 
                                common_area=excluded.common_area, 
                                total_area=excluded.total_area,
                                bank_area=excluded.bank_area
                        """,
                            (
                                row["asset_name"],
                                str(row["floor"]),
                                float(row["exclusive_area"]),
                                float(row["common_area"]),
                                float(row["total_area"]),
                                bank_area,
                            ),
                        )
                db_conn.commit()
                db_conn.close()
                fetch_data.clear()
                st.success("✅ 파일 업로드 및 DB 적용이 완료되었습니다.")
        except Exception as e:
            st.error(f"업로드 중 오류 발생: {e}")

    st.markdown("---")
    st.markdown("### 2. 자산 등록 및 수정")
    asset_update_mode = st.radio("작업 선택", ["✨ 신규 자산 등록", "📝 기존 자산 수정"], horizontal=True)

    sel_asset = ""
    sel_floor = ""
    default_exc = 0.0
    default_com = 0.0
    default_tot = 0.0
    default_bank = 0.0

    if asset_update_mode == "📝 기존 자산 수정":
        df_assets = fetch_data("SELECT * FROM Asset_Area")
        if df_assets.empty:
            st.warning("등록된 자산이 없습니다.")
        else:
            assets_list = df_assets["asset_name"].unique().tolist()
            sel_asset = st.selectbox("수정할 자산명 선택", assets_list)
            
            floors_list = df_assets[df_assets["asset_name"] == sel_asset]["floor"].unique().tolist()
            sel_floor = st.selectbox("수정할 층 선택", floors_list)
            
            if sel_asset and sel_floor:
                row = df_assets[(df_assets["asset_name"] == sel_asset) & (df_assets["floor"] == sel_floor)].iloc[0]
                default_exc = float(row.get("exclusive_area", 0.0))
                default_com = float(row.get("common_area", 0.0))
                default_tot = float(row.get("total_area", 0.0))
                default_bank = float(row.get("bank_area", 0.0))

    with st.form("asset_manual_form"):
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            if asset_update_mode == "✨ 신규 자산 등록":
                m_asset_name = st.text_input("자산명 (건물명)")
                m_floor = st.text_input("층수 (예: 1F, B1)")
            else:
                m_asset_name = st.text_input("자산명 (건물명)", value=sel_asset, disabled=True)
                m_floor = st.text_input("층수 (예: 1F, B1)", value=sel_floor, disabled=True)
                
        with col_m2:
            m_exclusive = st.number_input("전용 면적 (평)", min_value=0.0, step=1.0, value=default_exc)
            m_common = st.number_input("공용 면적 (평)", min_value=0.0, step=1.0, value=default_com)
            m_total = st.number_input("총 면적 (평)", min_value=0.0, step=1.0, value=default_tot)
            m_bank = st.number_input(
                "은행 및 지점 사용 면적 (평)", min_value=0.0, step=1.0, value=default_bank
            )

        submitted = st.form_submit_button("✅ 자산 정보 저장")
        if submitted:
            if m_asset_name and m_floor:
                try:
                    db_conn = engine.raw_connection()
                    c = db_conn.cursor()
                    c.execute(
                        """
                        INSERT INTO Asset_Area (asset_name, floor, exclusive_area, common_area, total_area, bank_area)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT(asset_name, floor) 
                        DO UPDATE SET 
                            exclusive_area=excluded.exclusive_area, 
                            common_area=excluded.common_area, 
                            total_area=excluded.total_area,
                            bank_area=excluded.bank_area
                    """,
                        (
                            m_asset_name,
                            m_floor,
                            float(m_exclusive),
                            float(m_common),
                            float(m_total),
                            float(m_bank),
                        ),
                    )
                    db_conn.commit()
                    db_conn.close()
                    fetch_data.clear()
                    st.success(
                        f"'{m_asset_name} {m_floor}' 정보가 성공적으로 저장되었습니다."
                    )
                except Exception as e:
                    st.error(f"저장 중 오류 발생: {e}")
            else:
                st.error("자산명과 층수는 필수 입력값입니다.")

    st.markdown("---")
    st.markdown("### 3. 등록된 자산 정보 삭제")
    st.info("※ 등록된 자산 정보를 완전히 삭제합니다 (오기입 수정 용도).")

    df_existing_assets = fetch_data("SELECT * FROM Asset_Area")
    if not df_existing_assets.empty:
        asset_options = df_existing_assets.apply(
            lambda x: f"{x['asset_name']} - {x['floor']}", axis=1
        ).tolist()
        selected_assets_to_delete = st.multiselect(
            "삭제할 자산/층 다중 선택", asset_options
        )

        if st.button("🗑️ 선택 자산 일괄 삭제", type="primary"):
            if not selected_assets_to_delete:
                st.warning("삭제할 대상을 선택해주세요.")
            else:
                try:
                    db_conn = engine.raw_connection()
                    c = db_conn.cursor()
                    for sel_item in selected_assets_to_delete:
                        sel_asset_name, sel_floor = sel_item.split(" - ", 1)
                        c.execute(
                            "DELETE FROM Asset_Area WHERE asset_name = %s AND floor = %s",
                            (sel_asset_name, sel_floor),
                        )
                    db_conn.commit()
                    db_conn.close()
                    fetch_data.clear()
                    st.success(
                        f"✅ {len(selected_assets_to_delete)}개의 자산 정보가 성공적으로 삭제되었습니다."
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"삭제 중 오류 발생: {e}")
    else:
        st.info("삭제할 자산 정보가 없습니다.")

# ==========================================
# Tab 5: 계약 업데이트
# ==========================================
with tab_contract_update:
    st.header("계약 등록 및 갱신/퇴점 처리")

    update_mode = st.radio(
        "작업 유형 선택",
        [
            "✨ 신규 계약",
            "🔄 계약 갱신",
            "📝 기존 계약 수정",
            "❌ 퇴점",
            "🗑️ 계약 완전 삭제",
            "📥 일괄 등록 (CSV/Excel)",
        ],
        horizontal=True,
    )

    df_asset_options = fetch_data("SELECT DISTINCT asset_name, floor FROM Asset_Area")
    asset_list = (
        df_asset_options["asset_name"].unique().tolist()
        if not df_asset_options.empty
        else []
    )

    if not asset_list:
        st.warning(
            "⚠️ 등록된 자산 정보가 없습니다. '자산정보 업데이트' 탭에서 자산을 먼저 등록해주세요."
        )
        st.stop()

    # Active contracts only for Renew/Terminate
    df_contracts_active = fetch_data(
        "SELECT * FROM Lease_Contracts WHERE status = 'ACTIVE' OR status IS NULL"
    )
    target_contract_id = None

    # ------------------
    # 1) 갱신/퇴점/삭제 모드일 경우 기존 계약 선택
    # ------------------
    if update_mode in ["🔄 계약 갱신", "📝 기존 계약 수정", "❌ 퇴점", "🗑️ 계약 완전 삭제"]:
        if update_mode == "🗑️ 계약 완전 삭제":
            df_for_selection = fetch_data("SELECT * FROM Lease_Contracts")
            warning_msg = "등록된 계약이 없습니다."
        else:
            df_for_selection = df_contracts_active
            warning_msg = "등록된 유효한(ACTIVE) 기존 계약이 없습니다."

        if df_for_selection.empty:
            st.warning(warning_msg)
            st.stop()

        options = df_for_selection.apply(
            lambda x: f"[{x['contract_id']}] {x['asset_name']} {x['floor']} - {x['company_name']} ({x['status']})",
            axis=1,
        ).tolist()

        if update_mode == "🗑️ 계약 완전 삭제":
            selected_contract_strs = st.multiselect(
                "삭제할 기존 계약 다중 선택", options
            )
            target_contract_ids = [
                int(sel.split("]")[0][1:]) for sel in selected_contract_strs
            ]
        else:
            selected_contract_str = st.selectbox("적용할 기존 계약 선택", options)
            target_contract_id = int(selected_contract_str.split("]")[0][1:])
            row_sel = df_for_selection[
                df_for_selection["contract_id"] == target_contract_id
            ].iloc[0]

    # ------------------
    # 2) 퇴점 폼 렌더링
    # ------------------
    if update_mode == "❌ 퇴점":
        st.markdown("---")
        st.markdown("#### 퇴점 처리 정보 입력")
        term_type = st.radio("퇴점 유형", ["만기 종료", "조기 종료"])

        col_t1, col_t2 = st.columns(2)
        with col_t1:
            term_date_default = pd.to_datetime(row_sel["end_date"]).date()
            if term_type == "조기 종료":
                term_date_default = datetime.now().date()
            new_end_date = st.date_input("최종 계약 종료일", value=term_date_default)
            deposit_return_date = st.date_input(
                "보증금 반환일", value=term_date_default
            )
        with col_t2:
            penalty_yn = st.selectbox("위약벌 여부", ["N", "Y"])
            penalty_amount = 0
            if penalty_yn == "Y":
                penalty_amount = st.number_input(
                    "위약벌(위약금) 청구 액수", min_value=0, step=1000000
                )

        if st.button(
            "❌ 선택 계약 퇴점 처리", type="primary", use_container_width=True
        ):
            if (
                new_end_date > pd.to_datetime(row_sel["end_date"]).date()
                and term_type == "조기 종료"
            ):
                st.error("조기 종료일은 기존 계약 종료일보다 늦을 수 없습니다.")
            else:
                try:
                    db_conn = engine.raw_connection()
                    c = db_conn.cursor()
                    # 기존 계약 상태 터미네이트로 변경 및 종료일 단축
                    c.execute(
                        """
                        UPDATE Lease_Contracts 
                        SET status = 'TERMINATED', end_date = %s, deposit_return_date = %s, penalty_yn = %s, penalty_amount = %s
                        WHERE contract_id = %s
                    """,
                        (
                            new_end_date.strftime("%Y-%m-%d"),
                            deposit_return_date.strftime("%Y-%m-%d"),
                            penalty_yn,
                            float(penalty_amount),
                            target_contract_id,
                        ),
                    )

                    # 이력 관리 추가
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    month_str = datetime.now().strftime("%Y-%m")
                    details_json = json.dumps(
                        {
                            "유형": term_type,
                            "종료일": new_end_date.strftime("%Y-%m-%d"),
                            "위약금": penalty_amount,
                        },
                        ensure_ascii=False,
                    )
                    c.execute(
                        """
                        INSERT INTO Contract_History (contract_id, action_type, action_date, action_month, details)
                        VALUES (%s, %s, %s, %s, %s)
                    """,
                        (
                            target_contract_id,
                            "퇴점",
                            today_str,
                            month_str,
                            details_json,
                        ),
                    )

                    db_conn.commit()
                    db_conn.close()
                    fetch_data.clear()
                    st.success(
                        "✅ 퇴점 처리가 완료되었습니다. 렌트롤은 조기종료년도까지만 반영되고 이후 목록에서 제외됩니다."
                    )
                except Exception as e:
                    st.error(f"오류 발생: {e}")

    # ------------------
    # 3) 삭제 폼 렌더링
    # ------------------
    elif update_mode == "🗑️ 계약 완전 삭제":
        st.markdown("---")
        st.markdown("#### 계약 일괄 완전 삭제")
        if not target_contract_ids:
            st.info("삭제할 계약을 1개 이상 선택해주세요.")
        else:
            st.warning(
                f"⚠️ 선택하신 {len(target_contract_ids)}개의 계약과 관련된 모든 이력 및 렌트롤 수동 조정 데이터가 완전히 삭제됩니다."
            )

            if st.button(
                "🗑️ 영구 일괄 삭제 진행", type="primary", use_container_width=True
            ):
                try:
                    db_conn = engine.raw_connection()
                    c = db_conn.cursor()
                    for t_id in target_contract_ids:
                        c.execute(
                            "DELETE FROM Lease_Contracts WHERE contract_id = %s",
                            (t_id,),
                        )
                        c.execute(
                            "DELETE FROM RentRoll_Overrides WHERE contract_id = %s",
                            (t_id,),
                        )
                        c.execute(
                            "DELETE FROM Contract_History WHERE contract_id = %s",
                            (t_id,),
                        )
                    db_conn.commit()
                    db_conn.close()
                    fetch_data.clear()
                    st.success(
                        f"✅ {len(target_contract_ids)}개의 계약이 데이터베이스에서 완전히 삭제되었습니다."
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"삭제 중 오류 발생: {e}")

    # ------------------
    # 4) 신규/갱신 폼 렌더링
    # ------------------
    elif update_mode in ["✨ 신규 계약", "🔄 계약 갱신", "📝 기존 계약 수정"]:
        default_vals = {
            "asset_name": asset_list[0],
            "floor": "",
            "company": "",
            "area": 0.0,
            "exclusive_area": 0.0,
            "c_date": datetime.now().date(),
            "s_date": datetime.now().date(),
            "e_date": datetime.now().date().replace(year=datetime.now().year + 2),
            "deposit": 0,
            "rent": 0,
            "maint": 0,
            "rf_details": [],
            "floor_details": {}
        }

        if update_mode in ["🔄 계약 갱신", "📝 기존 계약 수정"]:
            default_vals["asset_name"] = row_sel["asset_name"]
            default_vals["floor"] = row_sel["floor"]
            default_vals["company"] = row_sel["company_name"]
            default_vals["area"] = (
                float(row_sel["contract_area"])
                if pd.notnull(row_sel["contract_area"])
                else 0.0
            )
            default_vals["exclusive_area"] = (
                float(row_sel.get("contract_exclusive_area", 0.0))
                if pd.notnull(row_sel.get("contract_exclusive_area", 0.0))
                else 0.0
            )
            
            old_start = pd.to_datetime(row_sel["start_date"]).date()
            old_end = pd.to_datetime(row_sel["end_date"]).date()
            
            if update_mode == "🔄 계약 갱신":
                default_vals["s_date"] = old_end + timedelta(days=1)
                default_vals["e_date"] = default_vals["s_date"].replace(
                    year=default_vals["s_date"].year + 2
                )
            else: # 📝 기존 계약 수정
                if "contract_date" in row_sel and pd.notnull(row_sel["contract_date"]):
                    default_vals["c_date"] = pd.to_datetime(row_sel["contract_date"]).date()
                else:
                    default_vals["c_date"] = old_start
                default_vals["s_date"] = old_start
                default_vals["e_date"] = old_end
                
                if row_sel.get("rent_free_details"):
                    try:
                        default_vals["rf_details"] = json.loads(row_sel["rent_free_details"])
                    except:
                        pass
                if row_sel.get("floor_details"):
                    try:
                        default_vals["floor_details"] = json.loads(row_sel["floor_details"])
                    except:
                        pass

            default_vals["deposit"] = int(row_sel["deposit"])
            default_vals["rent"] = int(row_sel["monthly_rent"])
            default_vals["maint"] = int(row_sel["monthly_maintenance_fee"])

        with st.container():
            st.markdown(
                "<div style='background-color: white; padding: 2rem; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);'>",
                unsafe_allow_html=True,
            )

            st.markdown("#### 기본 계약 형태")
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                idx_ct = 0
                if update_mode in ["🔄 계약 갱신", "📝 기존 계약 수정"]:
                    if len(default_vals.get("floor_details", {})) > 1 or "," in default_vals["floor"]:
                        idx_ct = 1
                contract_type = st.radio(
                    "계약 형태", ["단층 계약", "복층 계약"], index=idx_ct, horizontal=True
                )
            with col_t2:
                idx_curr = 0
                if update_mode in ["🔄 계약 갱신", "📝 기존 계약 수정"]:
                    if "USD" in str(row_sel.get("currency", "")):
                        idx_curr = 1
                currency = st.radio("계약 통화", ["KRW", "USD"], index=idx_curr, horizontal=True)

            st.markdown("---")
            col_a, col_b = st.columns(2)
            with col_a:
                try:
                    asset_idx = asset_list.index(default_vals["asset_name"])
                except:
                    asset_idx = 0

                if update_mode == "🔄 계약 갱신":
                    asset_name = default_vals["asset_name"]
                    st.text_input(
                        "자산명 (갱신 시 고정)", value=asset_name, disabled=True
                    )
                else:
                    asset_name = st.selectbox(
                        "자산명 (기존 등록 자산)", asset_list, index=asset_idx
                    )

            with col_b:
                floor_list = df_asset_options[
                    df_asset_options["asset_name"] == asset_name
                ]["floor"].tolist()
                if contract_type == "단층 계약":
                    if update_mode == "🔄 계약 갱신":
                        floor_val = default_vals["floor"]
                        st.text_input("층 (고정)", value=floor_val, disabled=True)
                        sel_floors = [floor_val]
                    else:
                        try:
                            floor_idx = floor_list.index(default_vals["floor"])
                        except:
                            floor_idx = 0
                        if floor_list:
                            floor_val = st.selectbox(
                                "해당 층", floor_list, index=floor_idx
                            )
                        else:
                            floor_val = st.selectbox("해당 층", ["없음"])
                        sel_floors = [floor_val] if floor_val != "없음" else []
                else:
                    if update_mode == "🔄 계약 갱신":
                        st.info("복층 갱신 시 기존 층 정보를 재선택해주세요.")
                    def_floors = []
                    if update_mode == "📝 기존 계약 수정" and default_vals.get("floor_details"):
                        def_floors = [f for f in default_vals["floor_details"].keys() if f in floor_list]
                    sel_floors = st.multiselect("해당 층 다중 선택", floor_list, default=def_floors)

            st.markdown("---")
            st.markdown("#### 업체 및 면적 정보")
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                company_name = st.text_input(
                    "🏢 업체명 (임차인)", value=default_vals["company"]
                )

            floor_areas = {}
            with col_c2:
                if contract_type == "단층 계약":
                    col_f1, col_f2 = st.columns(2)
                    with col_f1:
                        contract_area = st.number_input(
                            "📐 계약 총면적 (평)",
                            min_value=0.0,
                            step=1.0,
                            value=default_vals["area"],
                        )
                    with col_f2:
                        contract_exclusive_area = st.number_input(
                            "📐 전용면적 (평)",
                            min_value=0.0,
                            step=1.0,
                            value=default_vals["exclusive_area"],
                        )
                    if sel_floors:
                        floor_areas[sel_floors[0]] = {
                            "area": contract_area,
                            "exclusive_area": contract_exclusive_area,
                        }
                else:
                    st.markdown("📐 **층별 계약 면적 (평)**")
                    contract_area = 0.0
                    contract_exclusive_area = 0.0
                    for fl in sel_floors:
                        def_fl_area = 0.0
                        def_fl_exc = 0.0
                        if update_mode == "📝 기존 계약 수정" and default_vals.get("floor_details"):
                            if fl in default_vals["floor_details"]:
                                def_fl_area = float(default_vals["floor_details"][fl].get("area", 0.0))
                                def_fl_exc = float(default_vals["floor_details"][fl].get("exclusive_area", 0.0))
                        col_f1, col_f2 = st.columns(2)
                        with col_f1:
                            fl_area = st.number_input(
                                f"{fl} 총면적",
                                min_value=0.0,
                                step=1.0,
                                value=def_fl_area,
                                key=f"area_{fl}",
                            )
                        with col_f2:
                            fl_exc_area = st.number_input(
                                f"{fl} 전용면적",
                                min_value=0.0,
                                step=1.0,
                                value=def_fl_exc,
                                key=f"exc_area_{fl}",
                            )
                        floor_areas[fl] = {
                            "area": fl_area,
                            "exclusive_area": fl_exc_area,
                        }
                        contract_area += fl_area
                        contract_exclusive_area += fl_exc_area
                    st.info(
                        f"총 면적 합계: {contract_area} 평 / 전용면적 합계: {contract_exclusive_area} 평"
                    )

            st.markdown("---")
            col_d1, col_d2, col_d3 = st.columns(3)
            with col_d1:
                contract_date = st.date_input(
                    "📝 새 계약 체결일", value=default_vals["c_date"]
                )
            with col_d2:
                start_date = st.date_input(
                    "🟢 새 임대 시작일", value=default_vals["s_date"]
                )
            with col_d3:
                end_date = st.date_input(
                    "🔴 새 임대 종료일", value=default_vals["e_date"]
                )

            st.markdown("---")
            st.markdown("#### 💳 임대 조건 (계약 전체 총액)")
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                dep_str = st.text_input(
                    "보증금 (총액)", value=f"{int(default_vals['deposit']):,}"
                )
                deposit = (
                    int(dep_str.replace(",", ""))
                    if dep_str.replace(",", "").isdigit()
                    else 0
                )
            with col_f2:
                rent_str = st.text_input(
                    "월 임대료 (총액)", value=f"{int(default_vals['rent']):,}"
                )
                monthly_rent = (
                    int(rent_str.replace(",", ""))
                    if rent_str.replace(",", "").isdigit()
                    else 0
                )
            with col_f3:
                maint_str = st.text_input(
                    "월 관리비 (총액)", value=f"{int(default_vals['maint']):,}"
                )
                monthly_maintenance_fee = (
                    int(maint_str.replace(",", ""))
                    if maint_str.replace(",", "").isdigit()
                    else 0
                )

            st.markdown("#### 📈 정기 인상 (Escalation)")
            use_escalation = st.checkbox("정기 인상 적용")
            escalation_cycle_years = 0
            rent_inc_rate = 0.0
            maint_inc_rate = 0.0
            if use_escalation:
                col_e1, col_e2, col_e3 = st.columns(3)
                with col_e1:
                    escalation_cycle_years = st.number_input(
                        "인상 주기 (년)", min_value=1, step=1, value=1
                    )
                with col_e2:
                    rent_inc_rate = st.number_input(
                        "임대료 인상률 (%)", min_value=0.0, step=0.1, value=5.0
                    )
                with col_e3:
                    maint_inc_rate = st.number_input(
                        "관리비 인상률 (%)", min_value=0.0, step=0.1, value=5.0
                    )

            st.markdown("---")
            st.markdown("#### 🎁 새로운 렌트프리 설정")
            available_months = get_months_between(start_date, end_date)
            valid_default_rf = [
                m for m in default_vals["rf_details"] if m in available_months
            ]
            selected_rf_months = st.multiselect(
                "렌트프리 적용 월 선택",
                options=available_months,
                default=valid_default_rf,
            )
            total_rf_months = len(selected_rf_months)

            st.markdown("---")
            st.markdown("#### ✍️ 비고 사항")
            remarks = st.text_area("특약 및 비고 사항을 입력하세요.")

            st.markdown("---")
            send_email = st.checkbox(
                "저장 완료 시 담당자에게 엑셀 보고서 발송", value=True
            )

            st.markdown("</div>", unsafe_allow_html=True)
            st.write("")

            if update_mode == "🔄 계약 갱신":
                try:
                    import proposal_generator
                    import importlib

                    importlib.reload(proposal_generator)
                    from proposal_generator import generate_renewal_proposal

                    old_data = {
                        "기존_총임대면적_평": row_sel["contract_area"],
                        "기존_전용면적_평": row_sel.get("contract_exclusive_area", 0),
                        "기존_월임대료": row_sel["monthly_rent"],
                        "기존_월관리비": row_sel["monthly_maintenance_fee"],
                        "기존_보증금": row_sel["deposit"],
                        "기존_임대차기간": f"{row_sel['start_date']} ~ {row_sel['end_date']}",
                    }
                    new_data = {
                        "자산주소": asset_name,
                        "GPMS_ID": f"C-{target_contract_id}",
                        "임차인명": company_name,
                        "부동산사용목적": "업무시설",
                        "대리인명": "",
                        "임대층": ", ".join(sel_floors) if sel_floors else "",
                        "신규_총임대면적_평": contract_area,
                        "신규_전용면적_평": contract_exclusive_area,
                        "갱신_보증금": deposit,
                        "갱신_월임대료": monthly_rent,
                        "갱신_월관리비": monthly_maintenance_fee,
                        "갱신_임대차기간": f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}",
                        "갱신_임대시작일": start_date.strftime("%Y-%m-%d"),
                        "갱신_임대만료일": end_date.strftime("%Y-%m-%d"),
                        "보증금비고": remarks,
                        "임대료비고": "",
                        "관리비비고": "",
                        "기간비고": "",
                    }

                    db_conn = engine.raw_connection()
                    try:
                        c = db_conn.cursor()
                        c.execute(
                            "SELECT floor, contract_area, deposit, monthly_rent, monthly_maintenance_fee FROM Lease_Contracts WHERE asset_name = %s AND status = 'ACTIVE' AND contract_id != %s",
                            (asset_name, target_contract_id),
                        )
                        comps_data = [
                            {
                                "floor": r[0],
                                "contract_area": r[1],
                                "deposit": r[2],
                                "monthly_rent": r[3],
                                "monthly_maintenance_fee": r[4],
                            }
                            for r in c.fetchall()
                        ]
                    finally:
                        db_conn.close()

                    file_bytes, filename = generate_renewal_proposal(
                        old_data, new_data, comps_data
                    )
                    st.download_button(
                        "📄 갱신 기안서류(Renewal Proposal) 자동생성",
                        data=file_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.warning(f"기안서류 생성 로딩 중 오류 (템플릿 확인 필요): {e}")

            if st.button(
                "✅ "
                + (
                    "신규 계약 등록 완료"
                    if update_mode == "✨ 신규 계약"
                    else "갱신 정보 저장 및 반영"
                ),
                use_container_width=True,
            ):
                if not asset_name or not company_name or not sel_floors:
                    st.error("자산명, 층, 업체명을 모두 올바르게 입력해주세요.")
                elif start_date > end_date:
                    st.error("종료일은 시작일보다 이후여야 합니다.")
                elif contract_area <= 0:
                    st.error("계약 면적이 0보다 커야 합니다.")
                else:
                    try:
                        db_conn = engine.raw_connection()
                        c = db_conn.cursor()
                        rf_details_json = json.dumps(selected_rf_months)
                        today_str = datetime.now().strftime("%Y-%m-%d")
                        month_str = datetime.now().strftime("%Y-%m")

                        floor_details_dict = {}
                        for fl, fl_info in floor_areas.items():
                            if isinstance(fl_info, dict):
                                fl_area = fl_info["area"]
                                fl_exc_area = fl_info["exclusive_area"]
                            else:
                                fl_area = fl_info
                                fl_exc_area = 0.0
                            floor_details_dict[fl] = {
                                "area": fl_area,
                                "exclusive_area": fl_exc_area,
                                "ratio": (
                                    fl_area / contract_area if contract_area > 0 else 0
                                ),
                            }
                        floor_details_json = json.dumps(floor_details_dict)

                        contract_months = len(get_months_between(start_date, end_date))
                        period_str = f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')} ({contract_months}개월)"
                        history_details = {
                            "계약기간": period_str,
                            "통화": currency,
                            "보증금": f"{int(deposit):,}",
                            "임대료": f"{int(monthly_rent):,}",
                            "관리비": f"{int(monthly_maintenance_fee):,}",
                        }

                        floor_str = ", ".join(sel_floors)

                        if update_mode == "✨ 신규 계약":
                            c.execute(
                                """
                                INSERT INTO Lease_Contracts (
                                    asset_name, floor, company_name, contract_date, start_date, end_date,
                                    contract_area, contract_exclusive_area, deposit, monthly_rent, monthly_maintenance_fee,
                                    total_rent_free_months, rent_free_details, status,
                                    currency, floor_details, escalation_cycle_years, rent_inc_rate, maint_inc_rate, remarks
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', %s, %s, %s, %s, %s, %s)
                            """,
                                (
                                    asset_name,
                                    floor_str,
                                    company_name,
                                    contract_date.strftime("%Y-%m-%d"),
                                    start_date.strftime("%Y-%m-%d"),
                                    end_date.strftime("%Y-%m-%d"),
                                    float(contract_area),
                                    float(contract_exclusive_area),
                                    float(deposit),
                                    float(monthly_rent),
                                    float(monthly_maintenance_fee),
                                    total_rf_months,
                                    rf_details_json,
                                    currency,
                                    floor_details_json,
                                    escalation_cycle_years,
                                    rent_inc_rate,
                                    maint_inc_rate,
                                    remarks,
                                ),
                            )
                            new_contract_id = c.lastrowid

                            c.execute(
                                """
                                INSERT INTO Contract_History (contract_id, action_type, action_date, action_month, details)
                                VALUES (%s, '신규', %s, %s, %s)
                            """,
                                (
                                    new_contract_id,
                                    today_str,
                                    month_str,
                                    json.dumps(history_details, ensure_ascii=False),
                                ),
                            )

                            db_conn.commit()
                            db_conn.close()
                            fetch_data.clear()
                            st.success(
                                f"🎉 '{company_name}' 신규 계약이 등록되었습니다."
                            )

                            if send_email:
                                try:
                                    wb = load_workbook("report_template.xlsx")
                                    ws = wb.active
                                    ws["B3"] = company_name
                                    ws["B4"] = asset_name
                                    ws["B5"] = floor_str
                                    ws["B6"] = start_date.strftime("%Y-%m-%d")
                                    ws["B7"] = end_date.strftime("%Y-%m-%d")
                                    ws["B8"] = deposit
                                    ws["B9"] = monthly_rent
                                    report_filename = f"report_{company_name}.xlsx"
                                    wb.save(report_filename)

                                    if "email" in st.secrets:
                                        msg = EmailMessage()
                                        msg["Subject"] = (
                                            f"[PM/AM] 신규 계약 체결 알림: {company_name}"
                                        )
                                        msg["From"] = st.secrets["email"]["user"]
                                        msg["To"] = st.secrets["email"]["receiver"]
                                        msg.set_content(
                                            f"신규 계약이 등록되었습니다.\n자산명: {asset_name}\n업체명: {company_name}"
                                        )

                                        with open(report_filename, "rb") as fa:
                                            msg.add_attachment(
                                                fa.read(),
                                                maintype="application",
                                                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                filename=report_filename,
                                            )

                                        with smtplib.SMTP_SSL(
                                            "smtp.gmail.com", 465
                                        ) as smtp:
                                            smtp.login(
                                                st.secrets["email"]["user"],
                                                st.secrets["email"]["password"],
                                            )
                                            smtp.send_message(msg)
                                        st.success(
                                            "📧 담당자에게 이메일 보고서가 발송되었습니다."
                                        )
                                    else:
                                        st.warning(
                                            "⚠️ .streamlit/secrets.toml 에 이메일 설정이 없어 메일 발송이 생략되었습니다."
                                        )
                                except Exception as email_err:
                                    st.warning(f"⚠️ 메일 발송 실패: {email_err}")

                        elif update_mode == "🔄 계약 갱신":
                            adjusted_old_end = start_date - timedelta(days=1)
                            c.execute(
                                """
                                UPDATE Lease_Contracts 
                                SET status = 'RENEWED', end_date = %s
                                WHERE contract_id = %s
                            """,
                                (
                                    adjusted_old_end.strftime("%Y-%m-%d"),
                                    target_contract_id,
                                ),
                            )

                            c.execute(
                                """
                                INSERT INTO Lease_Contracts (
                                    asset_name, floor, company_name, contract_date, start_date, end_date,
                                    contract_area, contract_exclusive_area, deposit, monthly_rent, monthly_maintenance_fee,
                                    total_rent_free_months, rent_free_details, status, parent_contract_id,
                                    currency, floor_details, escalation_cycle_years, rent_inc_rate, maint_inc_rate, remarks
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', %s, %s, %s, %s, %s, %s, %s)
                            """,
                                (
                                    asset_name,
                                    floor_str,
                                    company_name,
                                    contract_date.strftime("%Y-%m-%d"),
                                    start_date.strftime("%Y-%m-%d"),
                                    end_date.strftime("%Y-%m-%d"),
                                    float(contract_area),
                                    float(contract_exclusive_area),
                                    float(deposit),
                                    float(monthly_rent),
                                    float(monthly_maintenance_fee),
                                    total_rf_months,
                                    rf_details_json,
                                    target_contract_id,
                                    currency,
                                    floor_details_json,
                                    escalation_cycle_years,
                                    rent_inc_rate,
                                    maint_inc_rate,
                                    remarks,
                                ),
                            )
                            new_contract_id = c.lastrowid

                            history_details["이전계약ID"] = target_contract_id
                            c.execute(
                                """
                                INSERT INTO Contract_History (contract_id, action_type, action_date, action_month, details)
                                VALUES (%s, '갱신', %s, %s, %s)
                            """,
                                (
                                    new_contract_id,
                                    today_str,
                                    month_str,
                                    json.dumps(history_details, ensure_ascii=False),
                                ),
                            )

                            db_conn.commit()
                            db_conn.close()
                            fetch_data.clear()
                            st.success(
                                f"🎉 '{company_name}' 계약이 성공적으로 갱신(버전 분리) 처리되었습니다."
                            )
                            
                        elif update_mode == "📝 기존 계약 수정":
                            c.execute(
                                """
                                UPDATE Lease_Contracts SET
                                    asset_name = %s, floor = %s, company_name = %s, contract_date = %s, start_date = %s, end_date = %s,
                                    contract_area = %s, contract_exclusive_area = %s, deposit = %s, monthly_rent = %s, monthly_maintenance_fee = %s,
                                    total_rent_free_months = %s, rent_free_details = %s,
                                    currency = %s, floor_details = %s, escalation_cycle_years = %s, rent_inc_rate = %s, maint_inc_rate = %s, remarks = %s
                                WHERE contract_id = %s
                            """,
                                (
                                    asset_name,
                                    floor_str,
                                    company_name,
                                    contract_date.strftime("%Y-%m-%d"),
                                    start_date.strftime("%Y-%m-%d"),
                                    end_date.strftime("%Y-%m-%d"),
                                    float(contract_area),
                                    float(contract_exclusive_area),
                                    float(deposit),
                                    float(monthly_rent),
                                    float(monthly_maintenance_fee),
                                    total_rf_months,
                                    rf_details_json,
                                    currency,
                                    floor_details_json,
                                    escalation_cycle_years,
                                    rent_inc_rate,
                                    maint_inc_rate,
                                    remarks,
                                    target_contract_id,
                                ),
                            )
                            c.execute(
                                """
                                INSERT INTO Contract_History (contract_id, action_type, action_date, action_month, details)
                                VALUES (%s, '정보수정', %s, %s, %s)
                            """,
                                (
                                    target_contract_id,
                                    today_str,
                                    month_str,
                                    json.dumps(history_details, ensure_ascii=False),
                                ),
                            )
                            db_conn.commit()
                            db_conn.close()
                            fetch_data.clear()
                            st.success(f"✏️ '{company_name}' 기존 계약 정보가 성공적으로 수정(UPDATE)되었습니다.")
                            st.rerun()

                    except Exception as e:
                        st.error(f"데이터베이스 저장 오류: {e}")

    # ------------------
    # 5) 일괄 업로드 폼 렌더링
    # ------------------
    elif update_mode == "📥 일괄 등록 (CSV/Excel)":
        st.markdown("---")
        st.markdown("#### 계약 정보 일괄 업로드")

        template_cols = [
            "자산명",
            "업체명",
            "층",
            "통화",
            "계약 시작일",
            "계약 종료일",
            "전체면적",
            "전용면적",
            "보증금",
            "임대료",
            "관리비",
            "임대료 인상률",
            "관리비 인상률",
            "인상 주기",
            "기타 특약",
        ]
        example_data = {
            "자산명": "Pohang (작성예시)",
            "업체명": "포항시의사회",
            "층": "3F",
            "통화": "KRW",
            "계약 시작일": "2020-01-01",
            "계약 종료일": "2025-12-31",
            "전체면적": "39.9",
            "전용면적": "37.0",
            "보증금": "50000000",
            "임대료": "45430",
            "관리비": "0",
            "임대료 인상률": "5",
            "관리비 인상률": "0",
            "인상 주기": "1",
            "기타 특약": "매년 11월 1일 인상",
        }
        df_template = pd.DataFrame([example_data], columns=template_cols)
        csv_template = df_template.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "📝 빈 양식 다운로드 (CSV)",
            data=csv_template,
            file_name="contract_upload_template.csv",
            mime="text/csv",
        )
        st.info(
            "※ 다운로드한 양식에 맞춰 데이터를 입력하신 후 업로드해주세요. 모든 계약은 '단층'을 기준으로 임시 등록되며, 복층 등 특수 조건은 등록 후 개별 수정바랍니다."
        )

        uploaded_file = st.file_uploader(
            "작성된 계약정보 파일 선택", type=["csv", "xlsx", "xls"]
        )
        if uploaded_file:
            try:
                if uploaded_file.name.endswith(".csv"):
                    try:
                        df_up = pd.read_csv(uploaded_file)
                    except UnicodeDecodeError:
                        uploaded_file.seek(0)
                        df_up = pd.read_csv(uploaded_file, encoding="cp949")
                else:
                    df_up = pd.read_excel(uploaded_file)

                required_cols = ["자산명", "업체명", "계약 시작일"]
                missing_cols = [
                    col for col in required_cols if col not in df_up.columns
                ]
                if missing_cols:
                    st.error(
                        f"❌ 업로드된 파일에 다음 필수 컬럼이 누락되었습니다: {', '.join(missing_cols)}"
                    )
                    st.info(
                        "💡 다운로드한 빈 양식을 그대로 사용하여 첫 번째 행(헤더)이 변경되지 않도록 주의해주세요."
                    )
                else:

                    num_cols_up = [
                        "전체면적",
                        "전용면적",
                        "보증금",
                        "임대료",
                        "관리비",
                        "임대료 인상률",
                        "관리비 인상률",
                        "인상 주기",
                    ]
                    for c in num_cols_up:
                        if c in df_up.columns:
                            if pd.api.types.is_string_dtype(
                                df_up[c]
                            ) or pd.api.types.is_object_dtype(df_up[c]):
                                df_up[c] = (
                                    df_up[c]
                                    .astype(str)
                                    .str.replace(r"[^\d\.\-]", "", regex=True)
                                )
                                df_up[c] = df_up[c].replace("", "0")
                            df_up[c] = pd.to_numeric(df_up[c], errors="coerce").fillna(
                                0.0
                            )

                    st.markdown("#### 업로드된 데이터 미리보기")
                    display_styled_table(
                        center_styler(df_up).format(
                            {
                                c: "{:,.2f}"
                                for c in ["전체면적", "전용면적"]
                                if c in df_up.columns
                            }
                        )
                    )

                    if st.button(
                        "✅ 데이터베이스에 최종 반영하기",
                        type="primary",
                        key="btn_bulk_contract_upload",
                    ):
                        db_conn = engine.raw_connection()
                        c = db_conn.cursor()
                        inserted_count = 0
                        today_str = datetime.now().strftime("%Y-%m-%d")
                        month_str = datetime.now().strftime("%Y-%m")

                        for _, row in df_up.iterrows():
                            if (
                                pd.isna(row.get("자산명"))
                                or pd.isna(row.get("업체명"))
                                or pd.isna(row.get("계약 시작일"))
                            ):
                                continue
                            if "(작성예시)" in str(row.get("자산명", "")):
                                continue

                            asset_name = str(row.get("자산명", "")).strip()
                            company_name = str(row.get("업체명", "")).strip()
                            floor_str = str(row.get("층", "")).strip()
                            currency = str(row.get("통화", "KRW")).strip()
                            if not currency or currency.lower() == "nan":
                                currency = "KRW"

                            try:
                                s_date = pd.to_datetime(
                                    row.get("계약 시작일")
                                ).strftime("%Y-%m-%d")
                                e_date = pd.to_datetime(
                                    row.get("계약 종료일")
                                ).strftime("%Y-%m-%d")
                            except:
                                st.warning(
                                    f"'{company_name}' 계약의 날짜 형식이 잘못되어 건너뜁니다."
                                )
                                continue

                            def safe_float(val):
                                if pd.isna(val) or str(val).strip() == "":
                                    return 0.0
                                if isinstance(val, str):
                                    import re

                                    val = re.sub(r"[^\d\.\-]", "", str(val))
                                try:
                                    return float(val)
                                except:
                                    return 0.0

                            c_area = safe_float(row.get("전체면적"))
                            e_area = safe_float(row.get("전용면적"))
                            dep = safe_float(row.get("보증금"))
                            rent = safe_float(row.get("임대료"))
                            maint = safe_float(row.get("관리비"))
                            r_inc = safe_float(row.get("임대료 인상률"))
                            m_inc = safe_float(row.get("관리비 인상률"))
                            esc_cycle = int(safe_float(row.get("인상 주기")))
                            remarks = (
                                str(row.get("기타 특약", ""))
                                if pd.notnull(row.get("기타 특약"))
                                else ""
                            )
                            if remarks.lower() == "nan":
                                remarks = ""

                            floor_details_dict = {
                                floor_str: {"area": c_area, "ratio": 1.0}
                            }
                            floor_details_json = json.dumps(
                                floor_details_dict, ensure_ascii=False
                            )

                            c.execute(
                                """
                                INSERT INTO Lease_Contracts (
                                    asset_name, floor, company_name, contract_date, start_date, end_date,
                                    contract_area, contract_exclusive_area, deposit, monthly_rent, monthly_maintenance_fee,
                                    total_rent_free_months, rent_free_details, status,
                                    currency, floor_details, escalation_cycle_years, rent_inc_rate, maint_inc_rate, remarks
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', %s, %s, %s, %s, %s, %s)
                            """,
                                (
                                    asset_name,
                                    floor_str,
                                    company_name,
                                    s_date,
                                    s_date,
                                    e_date,
                                    c_area,
                                    e_area,
                                    dep,
                                    rent,
                                    maint,
                                    0,
                                    "[]",
                                    currency,
                                    floor_details_json,
                                    esc_cycle,
                                    r_inc,
                                    m_inc,
                                    remarks,
                                ),
                            )

                            new_id = c.lastrowid
                            history_details = {
                                "계약기간": f"{s_date} ~ {e_date}",
                                "통화": currency,
                                "보증금": f"{int(dep):,}",
                                "임대료": f"{int(rent):,}",
                                "관리비": f"{int(maint):,}",
                            }
                            c.execute(
                                """
                                INSERT INTO Contract_History (contract_id, action_type, action_date, action_month, details)
                                VALUES (%s, '신규(일괄등록)', %s, %s, %s)
                            """,
                                (
                                    new_id,
                                    today_str,
                                    month_str,
                                    json.dumps(history_details, ensure_ascii=False),
                                ),
                            )

                            inserted_count += 1

                        db_conn.commit()
                        db_conn.close()
                        fetch_data.clear()
                        if inserted_count > 0:
                            st.success(
                                f"✅ {inserted_count}건의 계약 정보 일괄 등록이 완료되었습니다."
                            )
                        else:
                            st.warning(
                                "등록된 데이터가 없습니다. 양식과 내용을 확인해주세요."
                            )
            except Exception as e:
                st.error(f"업로드 중 오류 발생: {e}")

# ==========================================
# Tab 6: 업데이트 이력 관리 (New Tab)
# ==========================================
with tab_history:
    st.header("계약 업데이트 이력 관리")
    st.markdown("신규, 갱신, 퇴점 등 모든 계약 변동 이력을 월별로 조회합니다.")

    # Check if History table has data
    df_history = fetch_data("""
        SELECT h.history_id, h.action_month, h.action_date, h.action_type, 
               c.asset_name, c.floor, c.company_name, h.details 
        FROM Contract_History h
        LEFT JOIN Lease_Contracts c ON h.contract_id = c.contract_id
        ORDER BY h.history_id DESC
    """)

    if not df_history.empty:

        def format_details(val):
            if not val:
                return ""
            try:
                parsed = json.loads(val)
                if isinstance(parsed, dict):
                    return " | ".join([f"{k}: {v}" for k, v in parsed.items()])
                return val
            except:
                return val

        df_history["details"] = df_history["details"].apply(format_details)

        df_history.rename(
            columns={
                "history_id": "이력ID",
                "action_month": "발생월",
                "action_date": "발생일",
                "action_type": "유형",
                "asset_name": "자산명",
                "floor": "층",
                "company_name": "업체명",
                "details": "상세내용",
            },
            inplace=True,
        )

        months = sorted(df_history["발생월"].unique().tolist(), reverse=True)
        types = sorted(df_history["유형"].unique().tolist())

        col_h1, col_h2 = st.columns(2)
        with col_h1:
            selected_month = st.selectbox("📅 이력 조회 연/월", ["전체보기"] + months)
        with col_h2:
            selected_types = st.multiselect(
                "🏷️ 이력 유형 필터 (전체보기 시 비워두세요)", options=types, default=[]
            )

        df_display_hist = df_history.copy()
        if selected_month != "전체보기":
            df_display_hist = df_display_hist[
                df_display_hist["발생월"] == selected_month
            ]
        if selected_types:
            df_display_hist = df_display_hist[
                df_display_hist["유형"].isin(selected_types)
            ]

        display_styled_table(center_styler(df_display_hist))

        renewals = df_display_hist[df_display_hist["유형"] == "갱신"]
        if not renewals.empty:
            st.markdown("---")
            st.markdown("#### 📄 갱신 이력 기안서류(Lease Renewal Proposal) 재출력")
            renewal_opts = renewals.apply(
                lambda x: f"[{x['이력ID']}] {x['자산명']} {x['층']} - {x['업체명']} ({x['발생일']})",
                axis=1,
            ).tolist()
            sel_hist_str = st.selectbox("다운로드할 갱신 이력 선택", renewal_opts)
            if sel_hist_str:
                hist_id = int(sel_hist_str.split("]")[0][1:])
                try:
                    db_conn = engine.raw_connection()
                    try:
                        c = db_conn.cursor()
                        new_contract_id = int(
                            c.execute(
                                "SELECT contract_id FROM Contract_History WHERE history_id = %s",
                                (hist_id,),
                            ).fetchone()[0]
                        )
                        details_str = c.execute(
                            "SELECT details FROM Contract_History WHERE history_id = %s",
                            (hist_id,),
                        ).fetchone()[0]
                        details_json = json.loads(details_str) if details_str else {}
                        old_contract_id = details_json.get("이전계약ID")
                    finally:
                        db_conn.close()
                    fetch_data.clear()

                    if old_contract_id:
                        new_c = fetch_data(
                            f"SELECT * FROM Lease_Contracts WHERE contract_id = {new_contract_id}"
                        ).iloc[0]
                        old_c = fetch_data(
                            f"SELECT * FROM Lease_Contracts WHERE contract_id = {old_contract_id}"
                        ).iloc[0]

                        import proposal_generator
                        import importlib

                        importlib.reload(proposal_generator)
                        from proposal_generator import generate_renewal_proposal

                        old_data = {
                            "기존_총임대면적_평": old_c["contract_area"],
                            "기존_전용면적_평": old_c.get("contract_exclusive_area", 0),
                            "기존_월임대료": old_c["monthly_rent"],
                            "기존_월관리비": old_c["monthly_maintenance_fee"],
                            "기존_보증금": old_c["deposit"],
                            "기존_임대차기간": f"{old_c['start_date']} ~ {old_c['end_date']}",
                        }
                        new_data = {
                            "자산주소": new_c["asset_name"],
                            "GPMS_ID": f"C-{old_contract_id}",
                            "임차인명": new_c["company_name"],
                            "부동산사용목적": "업무시설",
                            "대리인명": "",
                            "임대층": new_c["floor"],
                            "신규_총임대면적_평": new_c["contract_area"],
                            "신규_전용면적_평": new_c.get("contract_exclusive_area", 0),
                            "갱신_보증금": new_c["deposit"],
                            "갱신_월임대료": new_c["monthly_rent"],
                            "갱신_월관리비": new_c["monthly_maintenance_fee"],
                            "갱신_임대차기간": f"{new_c['start_date']} ~ {new_c['end_date']}",
                            "갱신_임대시작일": new_c["start_date"],
                            "갱신_임대만료일": new_c["end_date"],
                            "보증금비고": new_c["remarks"] if new_c["remarks"] else "",
                            "임대료비고": "",
                            "관리비비고": "",
                            "기간비고": "",
                        }

                        c.execute(
                            "SELECT floor, contract_area, deposit, monthly_rent, monthly_maintenance_fee FROM Lease_Contracts WHERE asset_name = %s AND status = 'ACTIVE' AND contract_id != %s",
                            (new_c["asset_name"], new_contract_id),
                        )
                        comps_data = [
                            {
                                "floor": r[0],
                                "contract_area": r[1],
                                "deposit": r[2],
                                "monthly_rent": r[3],
                                "monthly_maintenance_fee": r[4],
                            }
                            for r in c.fetchall()
                        ]

                        file_bytes, filename = generate_renewal_proposal(
                            old_data, new_data, comps_data
                        )
                        
                        col_dl1, col_dl2, col_dl3 = st.columns([5, 3, 2])
                        with col_dl1:
                            st.download_button(
                                "📥 선택한 이력 기안파일 다운로드",
                                data=file_bytes,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                            
                        with col_dl2:
                            to_email_6 = st.text_input("이메일", label_visibility="collapsed", placeholder="수신자 이메일 주소 입력", key=f"email_tab6_{hist_id}")
                            
                        with col_dl3:
                            if st.button("🚀 메일 발송", key=f"btn_email_tab6_{hist_id}", use_container_width=True):
                                if to_email_6:
                                    company_name = new_data.get('임차인명', '업체')
                                    success, err = send_email_with_attachment(
                                        to_email=to_email_6,
                                        subject=f"[PM/AM] {company_name} 갱신 기안서류",
                                        body=f"요청하신 {company_name}의 갱신 기안서류(Excel)를 첨부하여 보내드립니다.",
                                        file_bytes=file_bytes,
                                        file_name=filename,
                                        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                                    if success:
                                        st.toast("메일이 성공적으로 발송되었습니다!", icon="✅")
                                    else:
                                        st.error(f"메일 발송 실패: {err}")
                                else:
                                    st.warning("이메일 주소를 입력해주세요.")
                    else:
                        st.warning(
                            "선택하신 이력에는 이전 계약 정보가 포함되어 있지 않습니다."
                        )
                except Exception as e:
                    st.error(f"파일 생성 오류: {e}")
    else:
        st.info("아직 등록된 업데이트 이력이 없습니다.")
