import streamlit as st
import pandas as pd
import plotly.express as px
import psycopg2
import psycopg2.extras
import calendar
from datetime import datetime, timedelta
import json
import io
import smtplib
from email.message import EmailMessage
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def generate_formatted_excel(df, subtotal_indices=None):
    if subtotal_indices is None:
        subtotal_indices = []
        
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        wb = writer.book
        ws = writer.sheets['Sheet1']
        
        # 스타일 정의
        header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        header_font = Font(name='맑은 고딕', bold=True, color="000000")
        subtotal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        subtotal_font = Font(name='맑은 고딕', bold=True)
        zebra_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        data_font = Font(name='맑은 고딕')
        
        center_align = Alignment(horizontal='center', vertical='center')
        
        vertical_border = Border(
            top=Side(border_style=None), 
            bottom=Side(border_style=None),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        double_border = Border(
            top=Side(style='double'), 
            bottom=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        bottom_double_border = Border(
            top=Side(border_style=None),
            bottom=Side(style='double'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        # 헤더 스타일 (Top 1 Row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            # 헤더는 아래쪽에 얇은 실선이 있는게 좋으나, 샘플대로라면 세로만 있을수도 있음. 
            # 일반적인 헤더 가독성을 위해 상단/하단도 실선 적용 (원래 prompt 기준)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            
        ws.freeze_panes = 'A2'
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 데이터 서식
        for row in range(2, max_row + 1):
            is_subtotal = (row - 2) in subtotal_indices
            is_last_row = (row == max_row)
            is_even = (row % 2 == 0)
            
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # 공통 설정
                cell.font = data_font
                cell.alignment = center_align
                
                # 조건부 테두리 및 배경색
                current_border = vertical_border
                
                if is_even and not is_subtotal:
                    cell.fill = zebra_fill
                    
                if is_subtotal:
                    cell.fill = subtotal_fill
                    cell.font = subtotal_font
                    current_border = double_border
                elif is_last_row:
                    current_border = bottom_double_border
                
                cell.border = current_border
                
                # 숫자 및 날짜 서식
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                if isinstance(cell.value, datetime) or isinstance(cell.value, pd.Timestamp):
                    cell.number_format = 'yyyy-mm-dd'
                    
        # 열 너비 자동 맞춤
        for col in range(1, max_col + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            length = sum(1.5 if ord(c) > 127 else 1 for c in line)
                            if length > max_length:
                                max_length = length
                except:
                    pass
            adjusted_width = max_length * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width + 2, 50)
            
    return output.getvalue()

def add_subtotal_rows(df, group_col):
    new_rows = []
    subtotal_indices = []
    current_index = 0
    
    for name, group in df.groupby(group_col, sort=False):
        for _, row in group.iterrows():
            new_rows.append(row.to_dict())
            current_index += 1
            
        subtotal = {}
        for col in df.columns:
            if col == group_col:
                subtotal[col] = f"[{name} 소계]"
            elif col in ["Contract_ID", "contract_id", "id", "ID", "연도", "월", "계약일", "계약시작일", "계약종료일"]:
                subtotal[col] = None
            elif pd.api.types.is_numeric_dtype(df[col]):
                subtotal[col] = group[col].sum()
            else:
                subtotal[col] = None
        new_rows.append(subtotal)
        subtotal_indices.append(current_index)
        current_index += 1
        
    return pd.DataFrame(new_rows), subtotal_indices


# Custom CSS for Premium Design
st.markdown(
    """
<style>
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Noto+Sans+KR:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
    
    /* 1. 폰트 크기 및 스타일 일관성 확보 */
    html, body, .stApp, p, span, div, li, td, th {
        font-family: 'Pretendard', 'Inter', 'Noto Sans KR', sans-serif;
        -webkit-font-smoothing: antialiased;
        font-size: 13px !important;
        color: #334155;
    }
    
    h1, h2, h3 {
        color: #1e293b;
        font-weight: 700 !important;
        letter-spacing: -0.5px;
        margin-bottom: 0.5rem !important;
        padding-bottom: 0.5rem !important;
    }
    
    h1 { font-size: 1.5rem !important; border-bottom: 2px solid #e2e8f0; }
    h2 { font-size: 1.25rem !important; }
    h3 { font-size: 1.1rem !important; }

    /* 숫자/금액/날짜 데이터 고정폭 스타일 */
    .stNumberInput input, 
    [data-testid="stMetricValue"] div,
    .custom-st-table td {
        font-family: 'Pretendard', 'Roboto Mono', monospace !important;
        font-variant-numeric: tabular-nums;
    }

    /* 2. 요소 간 여백(Margin/Padding) 최소화 */
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 1.5rem !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
        max-width: 98% !important; 
    }
    
    .stApp { background-color: #F8FAFC; }

    /* Streamlit 기본 갭(Gap) 강제 축소 */
    [data-testid="stVerticalBlock"] > div {
        gap: 0.5rem !important;
    }
    [data-testid="stHorizontalBlock"] {
        gap: 0.5rem !important;
    }
    
    /* Buttons */
    .stButton>button {
        background: #ffffff;
        color: #334155 !important;
        border: 1px solid #CBD5E1;
        border-radius: 4px;
        padding: 0.2rem 0.5rem !important;
        font-weight: 600;
        font-size: 12px !important;
        box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
        height: 32px !important;
        line-height: 1 !important;
        min-height: 32px !important;
    }
    
    .stButton>button:hover {
        border-color: #005EB8;
        color: #005EB8 !important;
        background: #F8FAFC;
    }

    /* Cards and Containers */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border: 1px solid #E2E8F0 !important;
        border-radius: 4px !important;
        padding: 0.75rem !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
    }
    
    /* Metrics - 컴팩트화 */
    [data-testid="metric-container"] {
        background-color: #ffffff;
        border: 1px solid #E2E8F0;
        border-radius: 4px;
        padding: 0.75rem !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05);
        min-height: 80px !important;
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
    }
    [data-testid="stMetricValue"] > div {
        color: #005EB8 !important;
        font-weight: 700 !important;
        font-size: 1.2rem !important;
        margin-top: 4px;
    }
    [data-testid="stMetricLabel"] > div {
        color: #64748b !important;
        font-size: 12px !important;
        white-space: nowrap !important;
    }
    
    .stAlert {
        border-radius: 4px;
        border: 1px solid #E2E8F0;
        background-color: #ffffff;
        padding: 0.5rem 1rem !important;
        box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background-color: transparent;
        padding: 0;
        border-bottom: 1px solid #E2E8F0;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 36px;
        padding-left: 12px;
        padding-right: 12px;
        white-space: nowrap !important;
        background-color: transparent;
        border-radius: 0;
        color: #64748B;
        font-weight: 500;
        font-size: 13px !important;
        border-bottom: 2px solid transparent;
        transition: all 0.2s ease;
        padding-top: 0 !important;
        padding-bottom: 0 !important;
    }
    
    .stTabs [aria-selected="true"] {
        color: #005EB8 !important;
        font-weight: 600;
        border-bottom: 2px solid #005EB8 !important;
    }

    /* 3. 엄격한 가로/세로 정렬(Alignment) 적용 */
    .stTextInput>div>div>input, 
    .stNumberInput>div>div>input, 
    .stSelectbox>div>div>div,
    .stDateInput>div>div>input {
        border: 1px solid #CBD5E1 !important;
        border-radius: 4px;
        transition: all 0.2s ease;
        box-shadow: none !important;
        height: 32px !important;
        min-height: 32px !important;
        line-height: 32px !important;
        font-size: 13px !important;
        padding: 0 8px !important;
    }
    .stTextArea>div>div>textarea {
        border: 1px solid #CBD5E1 !important;
        border-radius: 4px;
        font-size: 13px !important;
        padding: 8px !important;
    }
    
    .stTextInput label, .stNumberInput label, .stSelectbox label, .stDateInput label, .stTextArea label {
        font-size: 12px !important;
        font-weight: 600 !important;
        margin-bottom: 2px !important;
        color: #475569 !important;
        min-height: 0 !important;
        padding-bottom: 0 !important;
    }

    .stTextInput>div>div>input:focus, 
    .stNumberInput>div>div>input:focus, 
    .stSelectbox>div>div>div:focus,
    .stTextArea>div>div>textarea:focus {
        border-color: #005EB8 !important;
        box-shadow: 0 0 0 1px #005EB8 !important;
    }

    .custom-st-table th:nth-child(1), .custom-st-table td:nth-child(1) {{
        position: -webkit-sticky;
        position: sticky;
        left: 0;
        z-index: 5;
        background-color: white;
        border-right: none !important;
    }}
    .custom-st-table th:nth-child(1) {
        z-index: 15;
        background-color: #F8FAFC !important;
        color: #334155 !important;
    }
    .custom-st-table tr:hover td:nth-child(1) {
        background-color: #f8fafc;
    }

    /* 절대 줄바꿈 금지 */
    .stButton>button, 
    [data-baseweb="tab"], 
    td, th, 
    [data-testid="stMetricLabel"], 
    [data-testid="stMetricValue"] {
        white-space: nowrap !important;
        word-break: keep-all !important;
        text-overflow: ellipsis;
    }
</style>
""",
    unsafe_allow_html=True,
)


# Helper function to send email with attachment
def send_email_with_attachment(to_email, subject, body, file_bytes, file_name, mime_type):
    if "email" not in st.secrets:
        return False, "Streamlit Cloud의 [Settings] -> [Secrets] 에 [email] 계정 정보를 입력해주세요! (자세한 설정 방법은 안내를 참고하세요)"
        
    try:
        user = st.secrets["email"]["user"]
        password = st.secrets["email"]["password"]
        host = st.secrets.get("email", {}).get("host", "smtp.gmail.com")
        port = st.secrets.get("email", {}).get("port", 465)
        
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = user
        msg["To"] = to_email
        msg.set_content(body)
        
        maintype, subtype = mime_type.split("/", 1)
        msg.add_attachment(file_bytes, maintype=maintype, subtype=subtype, filename=file_name)
        
        if port == 465:
            with smtplib.SMTP_SSL(host, port) as server:
                server.login(user, password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port) as server:
                server.starttls()
                server.login(user, password)
                server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)



# Constants for currency conversion (should be updated via API in production)
CURRENCY_RATES = {
    "USD_TO_KRW": 1400.0,  # TODO: Implement real-time exchange rate API
    "PY_TO_SQM": 3.3058,
    "PY_TO_SF": 35.583
}

# Performance: Cache rent schedule parsing
@st.cache_data(ttl=1800, show_spinner=False)
def _parse_rent_schedule(rent_schedule_json):
    """Parse and sort rent schedule once"""
    import json
    if rent_schedule_json:
        try:
            schedule = json.loads(rent_schedule_json)
            schedule.sort(key=lambda x: pd.to_datetime(x["start_date"]))
            return schedule
        except:
            pass
    return []


def get_scheduled_amount(rent_schedule_json, target_date, default_rent, default_maint, currency="KRW"):
    """Optimized: Use cached parsed schedule"""
    schedule = _parse_rent_schedule(rent_schedule_json)
    
    if not schedule:
        return default_rent, default_maint
    
    t_date = target_date.date()
    
    last_known_rent = default_rent
    last_known_maint = default_maint
    
    for period in schedule:
        try:
            s_dt = pd.to_datetime(period.get("start_date"))
            e_dt = pd.to_datetime(period.get("end_date"))
            if pd.isna(s_dt) or pd.isna(e_dt):
                continue
            s_date = s_dt.date()
            e_date = e_dt.date()
        except Exception:
            continue
        
        if t_date < s_date and last_known_rent == default_rent:
            return default_rent, default_maint
            
        if s_date <= t_date <= e_date:
            return float(period.get("rent", 0.0)), float(period.get("maint", 0.0))
        
        if t_date > e_date:
            last_known_rent = float(period.get("rent", 0.0))
            last_known_maint = float(period.get("maint", 0.0))
    
    return last_known_rent, last_known_maint


def highlight_total_row(row):
    is_total = False
    for val in row:
        if str(val).strip() == "TOTAL":
            is_total = True
            break
    if is_total:
        return ["background-color: #F8FAFC; font-weight: bold; color: #0F172A"] * len(
            row
        )
    return [""] * len(row)


# Helper function to center align dataframe and style headers/totals
def center_styler(df):
    # CSS is already handled globally in display_styled_table wrapper.
    # Removing set_properties and set_table_styles massively speeds up styler.to_html()
    styler = df.style.apply(highlight_total_row, axis=1)
    return styler


def format_money(x):
    if pd.isna(x):
        return ""
    try:
        val = float(x)
        if abs(val) >= 1_000_000:
            return f"{val/1_000_000:,.1f}백만"
        return f"{val:,.0f}"
    except:
        return str(x)

# Helper function to display styled table as HTML with scrolling
def display_styled_table(df, freeze_cols=1, format_dict=None, custom_css=""):
    import uuid
    import streamlit.components.v1 as components

    if hasattr(df, "data"):
        df = df.data

    auto_format = {}
    for col in df.columns:
        col_str = str(col)
        if any(k in col_str for k in ["면적", "비율", "율", "비중", "수익률"]):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = "{:,.2f}"
        elif any(
            k in col_str
            for k in [
                "금액",
                "보증금",
                "임대료",
                "관리비",
                "수익",
                "비용",
                "단가",
                "NOC",
                "월임대료",
                "월관리비",
                "합계",
            ]
        ):
            if pd.api.types.is_numeric_dtype(df[col]):
                auto_format[col] = format_money

    if format_dict:
        auto_format.update(format_dict)

    styler = df.style.apply(highlight_total_row, axis=1)
    if auto_format:
        styler = styler.format(auto_format)
    try:
        styler = styler.hide(axis="index")
    except Exception:
        try:
            styler = styler.hide_index()
        except:
            pass

    uid = "tbl_" + uuid.uuid4().hex[:8]
    html = styler.to_html()

    if freeze_cols == 4:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; min-width: 150px; max-width: 150px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}
        .{uid} th:nth-child(2), .{uid} td:nth-child(2) {{ position: -webkit-sticky; position: sticky; left: 150px; z-index: 5; min-width: 100px; max-width: 100px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}
        .{uid} th:nth-child(3), .{uid} td:nth-child(3) {{ position: -webkit-sticky; position: sticky; left: 250px; z-index: 5; min-width: 300px; max-width: 300px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .{uid} th:nth-child(4), .{uid} td:nth-child(4) {{ position: -webkit-sticky; position: sticky; left: 550px; z-index: 5; min-width: 60px; max-width: 60px; border-right: 1px solid #E2E8F0 !important; text-align: center !important; }}

        .{uid} th:nth-child(-n+4) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; text-align: center !important; }}
        """
    else:
        freeze_css = f"""
        .{uid} th:nth-child(1), .{uid} td:nth-child(1) {{ position: -webkit-sticky; position: sticky; left: 0; z-index: 5; border-right: 1px solid #E2E8F0 !important; }}
        .{uid} th:nth-child(1) {{ z-index: 15; background-color: #F8FAFC !important; color: #334155 !important; font-weight: 600 !important; }}
        """

    wrapper = f"""
<html>
<head>
<style>
body {{ margin: 0; font-family: 'Pretendard', 'Inter', sans-serif; -webkit-font-smoothing: antialiased; }}
.custom-st-table.{uid} {{
    width: 100%;
    max-width: 100%;
    overflow-x: auto;
    max-height: 550px;
    overflow-y: auto;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}}
.custom-st-table.{uid} table {{
    width: 100%;
    border-collapse: collapse;
    background-color: white;
    font-size: 13px;
    font-family: 'Pretendard', 'Roboto Mono', monospace;
    font-variant-numeric: tabular-nums;
    white-space: nowrap;
}}
.custom-st-table.{uid} th {{
    background-color: #F8FAFC !important;
    color: #334155 !important;
    font-weight: 600 !important;
    text-align: center !important;
    padding: 0.4rem 0.5rem !important;
    border-bottom: 1px solid #E2E8F0 !important;
    border-right: 1px solid #E2E8F0 !important;
    position: -webkit-sticky;
    position: sticky;
    top: 0;
    z-index: 10;
}}
.custom-st-table.{uid} td {{
    padding: 0.4rem 0.5rem !important;
    text-align: center !important;
    border-bottom: 1px solid #E2E8F0 !important;
    border-right: 1px solid #E2E8F0 !important;
    color: #334155;
}}
.custom-st-table.{uid} td:first-child {{
    text-align: center !important;
}}
.custom-st-table.{uid} td:last-child, .custom-st-table.{uid} th:last-child {{
    border-right: none !important;
}}
.custom-st-table.{uid} tr:nth-child(odd) td {{
    background-color: #ffffff;
}}
.custom-st-table.{uid} tr:nth-child(even) td {{
    background-color: #F8F9FA;
}}
.custom-st-table.{uid} tr:nth-child(odd) td:nth-child(even) {{
    background-color: #F1F3F5;
}}
.custom-st-table.{uid} tr:nth-child(even) td:nth-child(even) {{
    background-color: #E9ECEF;
}}
.custom-st-table.{uid} tr:hover td {{
    background-color: #E2E8F0 !important;
}}
{freeze_css}
{custom_css.replace('{uid}', uid)}
</style>
</head>
<body>
<div class="custom-st-table {uid}">
    {html}
</div>
</body>
</html>
"""
    components.html(wrapper, height=560, scrolling=False)


# DB Init
@st.cache_resource
def init_db():
    # Use Streamlit secrets for Supabase connection
    db_url = st.secrets["DATABASE_URL"]
    conn = psycopg2.connect(db_url)
    conn.autocommit = False
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS Asset_Area (
            asset_name TEXT,
            floor TEXT,
            exclusive_area REAL,
            common_area REAL,
            total_area REAL,
            PRIMARY KEY (asset_name, floor)
        )
    """)
    conn.commit()
    try:
        c.execute("ALTER TABLE Asset_Area ADD COLUMN bank_area REAL DEFAULT 0.0")
        conn.commit()
    except psycopg2.DatabaseError:
        conn.rollback()
        pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS Lease_Contracts (
            contract_id SERIAL PRIMARY KEY,
            asset_name TEXT,
            floor TEXT,
            company_name TEXT,
            contract_date DATE,
            start_date DATE,
            end_date DATE,
            contract_area REAL,
            deposit REAL,
            monthly_rent REAL,
            monthly_maintenance_fee REAL,
            total_rent_free_months INTEGER,
            rent_free_details TEXT
        )
    """)

    # Alter Lease_Contracts to add new columns safely
    new_columns = [
        ("status", "TEXT DEFAULT 'ACTIVE'"),
        ("deposit_return_date", "DATE"),
        ("penalty_yn", "TEXT"),
        ("penalty_amount", "REAL"),
        ("parent_contract_id", "INTEGER"),
        ("currency", "TEXT DEFAULT 'KRW'"),
        ("floor_details", "TEXT"),
        ("escalation_cycle_years", "INTEGER"),
        ("rent_inc_rate", "REAL"),
        ("maint_inc_rate", "REAL"),
        ("contract_exclusive_area", "REAL"),
        ("rent_schedule", "TEXT"),
        ("remarks", "TEXT"),
    ]
    for col_name, col_type in new_columns:
        try:
            c.execute(f"ALTER TABLE Lease_Contracts ADD COLUMN {col_name} {col_type}")
            if col_name == "status":
                c.execute(
                    "UPDATE Lease_Contracts SET status = 'ACTIVE' WHERE status IS NULL"
                )
            conn.commit()
        except psycopg2.DatabaseError:
            conn.rollback()
            pass  # Column already exists

    conn.commit()

    c.execute("DROP TABLE IF EXISTS RentRoll_Overrides")
    c.execute("""
        CREATE TABLE IF NOT EXISTS RentRoll_Overrides (
            contract_id INTEGER,
            floor TEXT,
            year INTEGER,
            month INTEGER,
            over_rent REAL,
            over_maint REAL,
            PRIMARY KEY (contract_id, floor, year, month)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS Contract_History (
            history_id SERIAL PRIMARY KEY,
            contract_id INTEGER,
            action_type TEXT,
            action_date DATE,
            action_month TEXT,
            details TEXT
        )
    """)
    conn.commit()
    return conn


conn = init_db()


@st.cache_resource
def get_engine():
    from sqlalchemy import create_engine

    return create_engine(
        st.secrets["DATABASE_URL"], pool_size=5, max_overflow=2, pool_recycle=300, pool_pre_ping=True
    )


engine = get_engine()


@st.cache_data(ttl=600, show_spinner=False)
def fetch_data(query, _eng=None):
    if _eng is None:
        _eng = get_engine()
    return pd.read_sql(query, _eng)


def get_floor_sort_key(floor_str):
    if not isinstance(floor_str, str):
        return -9999
    f = floor_str.upper().replace("F", "").strip()
    if f.startswith("B"):
        try:
            return -int(f[1:])
        except:
            return -9999
    else:
        try:
            return int(f)
        except:
            return 0


def sort_df_by_asset_and_floor(df, asset_col="asset_name", floor_col="floor"):
    import pandas as pd

    asset_df = fetch_data("SELECT asset_name, total_area FROM Asset_Area")
    if not asset_df.empty:
        sorted_assets = (
            asset_df.groupby("asset_name")["total_area"]
            .sum()
            .sort_values(ascending=False)
            .index.tolist()
        )
    else:
        sorted_assets = []

    if asset_col in df.columns:
        df[asset_col] = pd.Categorical(
            df[asset_col], categories=sorted_assets, ordered=True
        )
        sort_cols = [asset_col]
        asc = [True]
    else:
        sort_cols = []
        asc = []

    if floor_col in df.columns:
        df["_floor_sort"] = df[floor_col].apply(get_floor_sort_key)
        sort_cols.append("_floor_sort")
        asc.append(False)

    if sort_cols:
        df = df.sort_values(sort_cols, ascending=asc)
        if "_floor_sort" in df.columns:
            df = df.drop(columns=["_floor_sort"])

    return df


st.title("🏢 상업용 부동산 자산관리 시스템 (PM/AM)")

with st.sidebar:
    st.header("🔔 D-180 만기 도래 알림 데스크")
    df_active = fetch_data(
        "SELECT asset_name, floor, company_name, end_date FROM Lease_Contracts WHERE status = 'ACTIVE'"
    )

    if not df_active.empty:
        df_active["end_date"] = pd.to_datetime(df_active["end_date"], errors="coerce")
        today = pd.to_datetime(datetime.now().date())
        df_active["d_day"] = (df_active["end_date"] - today).dt.days

        # Filter 180 days or less, and >= 0 (not expired yet, or we can show expired as well)
        df_expiring = df_active[
            (df_active["d_day"] <= 180) & (df_active["d_day"] >= 0)
        ].sort_values("d_day")

        if not df_expiring.empty:
            for _, row in df_expiring.iterrows():
                if row["d_day"] <= 30:
                    st.error(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
                else:
                    st.warning(
                        f"[{row['asset_name']}-{row['floor']}] {row['company_name']} (D-{row['d_day']}일)"
                    )
        else:
            st.info("현재 6개월 내 만기 도래 계약이 없습니다.")
    else:
        st.info("현재 활성 계약이 없습니다.")


(
    tab_master_dashboard,
    tab_market_research,
    tab_asset_view,
    tab_stacking_plan,
    tab_lease_info,
    tab_rent_roll,
    tab_rent_change,
    tab_asset_update,
    tab_contract_update,
    tab_history,
) = st.tabs(
    [
        "🌐 마스터 대시보드",
        "📈 시장 동향 리서치",
        "📊 자산별 면적 현황",
        "🏢 스태킹 플랜",
        "📝 자산별 임대정보",
        "💰 렌트롤 (Rent Roll)",
        "📉 임관리비 변동 현황",
        "✏️ 자산정보 업데이트",
        "✍️ 계약 업데이트",
        "🕒 업데이트 이력 관리",
    ]
)


def get_months_between(start_date, end_date):
    months = []
    if not start_date or not end_date or start_date > end_date:
        return months
    current = datetime(start_date.year, start_date.month, 1)
    end = datetime(end_date.year, end_date.month, 1)
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    return months



@st.cache_data(ttl=86400)
def fetch_market_research_data():
    import random
    import pandas as pd
    import requests
    import streamlit as st
    
    # ---------------------------------------------------------
    # [실제 API 연동부] 
    # API 키는 .streamlit/secrets.toml의 R_ONE_API_KEY를 자동 참조합니다.
    # 추후 실제 Endpoint URL과 파라미터 구조가 확정되면 아래 주석을 풀고 연동합니다.
    # ---------------------------------------------------------
    api_key = st.secrets.get("R_ONE_API_KEY", None)
    api_endpoint = "https://api.reb.or.kr/v1/market/rent" # TODO: 실제 URL로 교체 필요
    
    if api_key and "TODO" not in api_endpoint:
        try:
            # 실제 연동 예시
            # response = requests.get(api_endpoint, params={"serviceKey": api_key, "format": "json"}, timeout=3)
            # response.raise_for_status()
            # items = response.json().get("response", {}).get("body", {}).get("items", [])
            # df = pd.DataFrame(items)
            # df["평당 임대료"] = (df["㎡당 임대료"] * 3.3058).round().astype(int)
            # return df
            pass
        except Exception as e:
            st.warning(f"API 연동 오류 (더미 데이터로 대체합니다): {e}")
            
    # ---------------------------------------------------------
    # API가 구성되지 않았거나 실패했을 때를 대비한 더미 데이터 생성 로직
    # ---------------------------------------------------------
    regions = ["서울", "경기", "인천", "부산", "대구", "광주", "대전"]
    sub_regions = {
        "서울": ["강남대로", "테헤란로", "도산대로", "여의도", "광화문", "명동", "홍대합정"],
        "경기": ["분당", "판교", "일산", "평촌"],
        "인천": ["부평", "구월", "송도"],
        "부산": ["서면", "해운대", "광복동"],
        "대구": ["동성로", "수성구"],
        "광주": ["상무지구", "충장로"],
        "대전": ["둔산", "은행동"]
    }
    asset_types = ["오피스", "소규모 상가", "중대형 상가"]
    quarters = ["2023 1Q", "2023 2Q", "2023 3Q", "2023 4Q", "2024 1Q", "2024 2Q"]
    
    data = []
    for r in regions:
        for sr in sub_regions[r]:
            for at in asset_types:
                for q in quarters:
                    base_rent = random.uniform(15000, 35000) if at == "오피스" else random.uniform(20000, 60000)
                    if r == "서울":
                        base_rent *= 1.5
                    vacancy = random.uniform(2.0, 15.0)
                    data.append({
                        "지역명(시/도)": r,
                        "세부 상권명": sr,
                        "자산 유형": at,
                        "기준 분기": q,
                        "㎡당 임대료": round(base_rent),
                        "공실률(%)": round(vacancy, 1)
                    })
    
    df = pd.DataFrame(data)
    df["평당 임대료"] = df["㎡당 임대료"] * 3.3058
    return df

# ==========================================
# Tab 0: 마스터 대시보드

# ==========================================
