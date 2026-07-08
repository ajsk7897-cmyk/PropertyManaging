import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime

def create_template():
    wb = Workbook()
    
    # Sheet 1
    ws1 = wb.active
    ws1.title = 'Template Frame 1'
    
    # Sheet 2
    ws2 = wb.create_sheet(title='Template Frame 2')
    
    # Fill some basic headers as a placeholder
    headers1 = ['임대갱신품의서', '항목', '내용', '비고']
    headers2 = ['계약조건 비교표', '기존 조건', '변경 조건', '증감', '비고']
    
    ws1.append(headers1)
    ws2.append(headers2)
    
    # Apply styles to both sheets
    fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    font = Font(bold=True)
    alignment = Alignment(vertical='center', horizontal='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border
            
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[column].width = adjusted_width

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H_%M')
    filename = f'부동산 임대차 자동생성 템플릿_{timestamp}.xlsx'
    wb.save(filename)
    print(f'Template saved as {filename}')

if __name__ == '__main__':
    create_template()
