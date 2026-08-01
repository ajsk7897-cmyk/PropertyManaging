import io
import math
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
from copy import copy

def py_to_sqm(py):
    if not py: return ""
    return math.floor(float(py) * 3.3058 * 100) / 100

def py_to_sf(py):
    if not py: return ""
    return math.floor(float(py) * 35.5832 * 100) / 100

def safe_str(val):
    if val is None or str(val).strip() == "" or str(val).lower() == "none":
        return ""
    return str(val)

def copy_cell_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font) if source_cell.font else None
    target_cell.border = copy(source_cell.border) if source_cell.border else None
    target_cell.fill = copy(source_cell.fill) if source_cell.fill else None
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection) if source_cell.protection else None
    target_cell.alignment = copy(source_cell.alignment) if source_cell.alignment else None

def generate_renewal_proposal(old_data, new_data, comps_list=None):
    template_path = '湲곗븞?뚯씪/Lease_Renewal_Proposal.xlsx'
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"?쒗뵆由??뚯씪??議댁옱?섏? ?딆뒿?덈떎: {template_path}")
        
    wb = load_workbook(template_path)
    
    def set_value(ws_name, cell, value, num_format=None):
        if isinstance(ws_name, int):
            try:
                ws = wb.worksheets[ws_name]
            except IndexError:
                return
        elif ws_name in wb.sheetnames:
            ws = wb[ws_name]
        else:
            return
            
        if value is None or str(value).lower() == "none":
            value = ""
        target_cell = None
        try:
                ws[cell].value = value
                target_cell = ws[cell]
            except AttributeError:
                for merged_range in ws.merged_cells.ranges:
                    if cell in merged_range:
                        top_left = merged_range.coord.split(':')[0]
                        ws[top_left].value = value
                        target_cell = ws[top_left]
                        break
            if target_cell and num_format:
                target_cell.number_format = num_format

    def get_money(val):
        if val is None or str(val).strip() == "" or str(val).lower() == "none":
            return 0
        try:
            return math.floor(float(val) / 10) * 10
        except (ValueError, TypeError):
            return 0

    old_gross_py = float(old_data.get('湲곗〈_珥앹엫?硫댁쟻_??) or 0)
    old_exc_py = float(old_data.get('湲곗〈_?꾩슜硫댁쟻_??) or 0)
    
    new_gross_py = float(new_data.get('?좉퇋_珥앹엫?硫댁쟻_??) or 0)
    new_exc_py = float(new_data.get('?좉퇋_?꾩슜硫댁쟻_??) or 0)

    # Date parsing logic for calculation
    new_start_str = safe_str(new_data.get('媛깆떊_?꾨??쒖옉??))
    new_end_str = safe_str(new_data.get('媛깆떊_?꾨?留뚮즺??))
    
    dt_start = None
    dt_end = None
    renewal_months = 0
    try:
        dt_start = datetime.datetime.strptime(new_start_str, "%Y-%m-%d").date()
        dt_end = datetime.datetime.strptime(new_end_str, "%Y-%m-%d").date()
        renewal_months = math.floor((dt_end - dt_start).days / 365 * 12)
    except Exception:
        pass

    # 媛. '?꾨?媛깆떊?덉쓽?? ?쒗듃
    set_value('?꾨?媛깆떊?덉쓽??, 'D6', safe_str(new_data.get('?먯궛二쇱냼')))
    set_value('?꾨?媛깆떊?덉쓽??, 'K6', safe_str(new_data.get('GPMS_ID')))
    set_value('?꾨?媛깆떊?덉쓽??, 'D7', safe_str(new_data.get('?꾩감?몃챸')))
    set_value('?꾨?媛깆떊?덉쓽??, 'K7', safe_str(new_data.get('遺?숈궛?ъ슜紐⑹쟻')))
    set_value('?꾨?媛깆떊?덉쓽??, 'D8', safe_str(new_data.get('?由ъ씤紐?)))
    set_value('?꾨?媛깆떊?덉쓽??, 'K8', safe_str(new_data.get('?꾨?痢?)))

    set_value('?꾨?媛깆떊?덉쓽??, 'D11', old_gross_py if old_gross_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'K11', old_exc_py if old_exc_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'D12', py_to_sqm(old_gross_py) if old_gross_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'K12', py_to_sqm(old_exc_py) if old_exc_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'D13', py_to_sf(old_gross_py) if old_gross_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'K13', py_to_sf(old_exc_py) if old_exc_py else "", num_format='#,##0.00')
    
    old_rent = get_money(old_data.get('湲곗〈_?붿엫?猷?))
    old_maint = get_money(old_data.get('湲곗〈_?붽?由щ퉬'))
    old_dep = get_money(old_data.get('湲곗〈_蹂댁쬆湲?))
    new_rent = get_money(new_data.get('媛깆떊_?붿엫?猷?))
    new_maint = get_money(new_data.get('媛깆떊_?붽?由щ퉬'))
    new_dep = get_money(new_data.get('媛깆떊_蹂댁쬆湲?))

    set_value('?꾨?媛깆떊?덉쓽??, 'D14', old_rent, num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'J14', old_maint, num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'D15', old_dep, num_format='#,##0')

    # [?섏젙 1] 怨꾩빟 媛쒖썡 ??(?덉궗???뺤닔)
    set_value('?꾨?媛깆떊?덉쓽??, 'D18', f"{renewal_months}媛쒖썡" if renewal_months > 0 else "")

    # [?섏젙 2] Date ?щ㎎ 蹂寃?(?묒? ?쒖떇)
    ws_ren = wb['?꾨?媛깆떊?덉쓽??]
    if dt_end:
        set_value('?꾨?媛깆떊?덉쓽??, 'D17', dt_end, num_format='[$-en-US]dd-mmm-yyyy;@')
        set_value('?꾨?媛깆떊?덉쓽??, 'J21', dt_end, num_format='[$-en-US]dd-mmm-yyyy;@')
    else:
        set_value('?꾨?媛깆떊?덉쓽??, 'D17', new_end_str)
        set_value('?꾨?媛깆떊?덉쓽??, 'J21', new_end_str)

    if dt_start:
        set_value('?꾨?媛깆떊?덉쓽??, 'D22', dt_start, num_format='[$-en-US]dd-mmm-yyyy;@')
    else:
        set_value('?꾨?媛깆떊?덉쓽??, 'D22', new_start_str)

    set_value('?꾨?媛깆떊?덉쓽??, 'D19', new_gross_py if new_gross_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'J18', new_exc_py if new_exc_py else "", num_format='#,##0.00')
    set_value('?꾨?媛깆떊?덉쓽??, 'J19', new_maint, num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'D20', new_rent, num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'D21', new_dep, num_format='#,##0')

    set_value('?꾨?媛깆떊?덉쓽??, 'A17', '??Renewal Proposal 媛깆떊 議곌굔')

    # [?섏젙 3] 怨좎젙 ?щТ 吏???섏떇 ?곸슜 (D25, J25, G27)
    set_value('?꾨?媛깆떊?덉쓽??, 'J22', '=(D21*J20)/12+D20+J19', num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'D25', '=(D15*J20)+(D14*12)+(J14*12)', num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'J25', '=(D21*J20)+(D20*12)+(J19*12)', num_format='#,##0')
    set_value('?꾨?媛깆떊?덉쓽??, 'G27', '=(D21*J20)/2+(D20*6)+(J19*6)', num_format='#,##0')

    # C34 : ?묒뾽???怨꾩빟???랁븳 ?먯궛??嫄대Ъ ?꾩껜???됯퇏 ?됰떦 愿由щ퉬
    total_maint_per_py = 0
    count_contracts = 0
    
    if old_gross_py > 0:
        total_maint_per_py += (old_maint / old_gross_py)
        count_contracts += 1
        
    if comps_list:
        for comp in comps_list:
            comp_area = float(comp.get('contract_area') or 0)
            if comp_area > 0:
                comp_maint = float(comp.get('monthly_maintenance_fee') or 0)
                total_maint_per_py += (comp_maint / comp_area)
                count_contracts += 1
                
    avg_maint_per_py = total_maint_per_py / count_contracts if count_contracts > 0 else 0
    set_value('?꾨?媛깆떊?덉쓽??, 'C34', get_money(avg_maint_per_py), num_format='#,##0')

    # A38 : 吏?뺣맂 ?띿뒪???낅젰
    a38_text = f"""[怨꾩빟 紐?

1. ?묒쓽 History
1) ?묒쓽 ?댁뒋?ы빆
 -  吏꾪뻾

2) ?꾨?李?議곌굔
 -   
  -> 蹂댁쬆湲?: 
  -> ?꾨?猷?: 
  -> 愿由щ퉬 : 
  -> Rent-Free : 
  -> 怨꾩빟湲곌컙 : 

2. 寃곕줎
 - """
    set_value('?꾨?媛깆떊?덉쓽??, 'A38', a38_text)

    # ?? '鍮꾧탳?? ?쒗듃
    set_value('鍮꾧탳??, 'A2', safe_str(new_data.get('?꾩감?몃챸')))
    set_value('鍮꾧탳??, 'J10', f"{renewal_months}媛쒖썡" if renewal_months > 0 else "")

    set_value('鍮꾧탳??, 'D4', py_to_sf(old_exc_py) if old_exc_py else "", num_format='#,##0.00')
    set_value('鍮꾧탳??, 'G4', "醫뚮룞" if old_exc_py == new_exc_py else py_to_sf(new_exc_py), num_format=None if old_exc_py == new_exc_py else '#,##0.00')
    
    set_value('鍮꾧탳??, 'D5', py_to_sf(old_gross_py) if old_gross_py else "", num_format='#,##0.00')
    set_value('鍮꾧탳??, 'G5', "醫뚮룞" if old_gross_py == new_gross_py else py_to_sf(new_gross_py), num_format=None if old_gross_py == new_gross_py else '#,##0.00')
    
    set_value('鍮꾧탳??, 'D6', safe_str(new_data.get('?꾩감?몃챸')))
    set_value('鍮꾧탳??, 'G6', "醫뚮룞")
    
    set_value('鍮꾧탳??, 'D7', old_dep, num_format='#,##0')
    set_value('鍮꾧탳??, 'G7', "醫뚮룞" if old_dep == new_dep else new_dep, num_format=None if old_dep == new_dep else '#,##0')
    set_value('鍮꾧탳??, 'K7', safe_str(new_data.get('蹂댁쬆湲덈퉬怨?)))
    
    rent_inc_str = ""
    if old_rent > 0 and old_rent != new_rent:
        inc_pct = ((new_rent - old_rent) / old_rent) * 100
        rent_inc_str = f"{inc_pct:.1f}% ?몄긽" if inc_pct > 0 else f"{abs(inc_pct):.1f}% ?명븯"
        
    set_value('鍮꾧탳??, 'D8', old_rent, num_format='#,##0')
    set_value('鍮꾧탳??, 'G8', "醫뚮룞" if old_rent == new_rent else new_rent, num_format=None if old_rent == new_rent else '#,##0')
    set_value('鍮꾧탳??, 'K8', rent_inc_str if rent_inc_str else safe_str(new_data.get('?꾨?猷뚮퉬怨?)))
    
    maint_inc_str = ""
    if old_maint > 0 and old_maint != new_maint:
        inc_pct = ((new_maint - old_maint) / old_maint) * 100
        maint_inc_str = f"{inc_pct:.1f}% ?몄긽" if inc_pct > 0 else f"{abs(inc_pct):.1f}% ?명븯"
        
    set_value('鍮꾧탳??, 'D9', old_maint, num_format='#,##0')
    set_value('鍮꾧탳??, 'G9', "醫뚮룞" if old_maint == new_maint else new_maint, num_format=None if old_maint == new_maint else '#,##0')
    set_value('鍮꾧탳??, 'K9', maint_inc_str if maint_inc_str else safe_str(new_data.get('愿由щ퉬鍮꾧퀬')))
    
    old_term = safe_str(old_data.get('湲곗〈_?꾨?李④린媛?))
    new_term = safe_str(new_data.get('媛깆떊_?꾨?李④린媛?))
    set_value('鍮꾧탳??, 'D10', old_term)
    set_value('鍮꾧탳??, 'G10', "醫뚮룞" if old_term == new_term else new_term)
    set_value('鍮꾧탳??, 'K10', "怨꾩빟媛깆떊")

    # [?섏젙 4, 5] 鍮꾧탳???곗감蹂??ㅽ뀦???숈쟻 ?뚮뜑留?    ws_comp = wb['鍮꾧탳??]
    step_ups = new_data.get('step_ups', [])
    years = max(2, math.ceil(renewal_months / 12) if renewal_months > 0 else 2)
    
    # Base columns for 1st, 2nd year: D=4, G=7 (spacing is 3)
    start_col_idx = 4
    col_spacing = 3

    for y in range(1, years + 1):
        target_col_idx = start_col_idx + (y - 1) * col_spacing
        target_col = get_column_letter(target_col_idx)
        
        # Determine values
        if y <= len(step_ups):
            y_rent = int(step_ups[y-1].get('rent') or new_rent)
            y_maint = int(step_ups[y-1].get('maint') or new_maint)
        else:
            y_rent = new_rent
            y_maint = new_maint
            
        y_rent_annual = y_rent * 12
        y_maint_annual = y_maint * 12

        # Write text and values
        set_value('鍮꾧탳??, f"{target_col}13", f"{y}??李?)
        set_value('鍮꾧탳??, f"{target_col}14", new_dep, num_format='#,##0')
        set_value('鍮꾧탳??, f"{target_col}15", y_rent_annual, num_format='#,##0')
        set_value('鍮꾧탳??, f"{target_col}16", y_maint_annual, num_format='#,##0')
        set_value('鍮꾧탳??, f"{target_col}17", f"={target_col}14*?꾨?媛깆떊?덉쓽??J20", num_format='#,##0')
        set_value('鍮꾧탳??, f"{target_col}18", f"=SUM({target_col}15:{target_col}17)", num_format='#,##0')
        
        # If y > 2, we must copy styles from 2nd year (col G)
        if y > 2:
            src_col = "G"
            # 13?됰???18?됯퉴吏 ?쒖떇 蹂듭궗, 蹂묓빀???댁빞?섏?留??⑥닚?붾? ?꾪빐 媛?? ?쒖떇 蹂듭궗
            for row in range(13, 19):
                src_cell = ws_comp[f"{src_col}{row}"]
                tgt_cell = ws_comp[f"{target_col}{row}"]
                copy_cell_style(src_cell, tgt_cell)

    # Move Total Column (J??濡쒖쭅) to the right
    total_col_idx = start_col_idx + years * col_spacing
    total_col = get_column_letter(total_col_idx)
    
    # 13??18???⑷퀎 ?쒖떇 洹몃━湲?    for row in range(13, 19):
        # 湲곗〈 J???쒖떇??蹂듭궗 (J??= 10)
        src_cell = ws_comp[f"J{row}"]
        tgt_cell = ws_comp[f"{total_col}{row}"]
        copy_cell_style(src_cell, tgt_cell)
            
        if row == 13:
            set_value('鍮꾧탳??, f"{total_col}{row}", "?⑷퀎(??")
        elif row == 14:
            set_value('鍮꾧탳??, f"{total_col}{row}", "") # 蹂댁쬆湲?怨듬?
        else:
            # SUM range: D15:M15 (for 3 years)
            end_sum_col = get_column_letter(total_col_idx - col_spacing)
            set_value('鍮꾧탳??, f"{total_col}{row}", f"=SUM(D{row}:{end_sum_col}{row})", num_format='#,##0')

    # [鍮꾧탳 ?щ?(Comps) ?됰떒媛 ?먮룞 異붿텧] - 遺媛???쒖쇅, 珥??꾨?硫댁쟻 湲곗?
    # C30: ?묒뾽 以묒씤 ?뚮꼳??紐?    set_value('?꾨?媛깆떊?덉쓽??, 'C30', safe_str(new_data.get('?꾩감?몃챸')))
    # C31 ~ C35 怨꾩궛
    if new_gross_py > 0:
        set_value('?꾨?媛깆떊?덉쓽??, 'C31', new_dep / new_gross_py, num_format='#,##0')
        set_value('?꾨?媛깆떊?덉쓽??, 'C32', new_rent / new_gross_py, num_format='#,##0')
        set_value('?꾨?媛깆떊?덉쓽??, 'C33', new_maint / new_gross_py, num_format='#,##0')
        set_value('?꾨?媛깆떊?덉쓽??, 'C35', (new_rent * 100 + new_dep) / new_gross_py, num_format='#,##0')
        
    for col in ['E', 'F', 'G', 'H', 'I', 'J']:
        set_value('?꾨?媛깆떊?덉쓽??, f'{col}35', "")

    if comps_list and len(comps_list) > 0 and new_gross_py > 0:
        target_rent_per_py = new_rent / new_gross_py
        
        comps_sorted = []
        for comp in comps_list:
            comp_area = float(comp.get('contract_area', 0))
            if comp_area <= 0: continue
            comp_rent = float(comp.get('monthly_rent', 0))
            comp_rent_per_py = comp_rent / comp_area
            diff = abs(comp_rent_per_py - target_rent_per_py)
            comps_sorted.append((diff, comp_rent_per_py, comp))
            
        comps_sorted.sort(key=lambda x: x[0])
        top_3_comps = comps_sorted[:3]
        
        cols = ['E', 'F', 'G']
        row_floor = 13     
        row_dep_py = 14    
        row_rent_py = 15   
        row_maint_py = 16  
        
        for idx, (diff, rent_per_py, comp) in enumerate(top_3_comps):
            if idx >= 3: break
            col = cols[idx]
            comp_area = float(comp.get('contract_area', 1))
            dep_per_py = float(comp.get('deposit', 0)) / comp_area
            maint_per_py = float(comp.get('monthly_maintenance_fee', 0)) / comp_area
            
            set_value('鍮꾧탳??, f'{col}{row_floor}', safe_str(comp.get('floor')))
            set_value('鍮꾧탳??, f'{col}{row_dep_py}', get_money(dep_per_py), num_format='#,##0')
            set_value('鍮꾧탳??, f'{col}{row_rent_py}', get_money(rent_per_py), num_format='#,##0')
            set_value('鍮꾧탳??, f'{col}{row_maint_py}', get_money(maint_per_py), num_format='#,##0')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    company_name = safe_str(new_data.get('?꾩감?몃챸'))
    if not company_name:
        company_name = "?꾩감??
    today_str = datetime.datetime.now().strftime('%Y-%m-%d')
    filename = f"DOA_{company_name}-{today_str}.xlsx"
    
    return output.getvalue(), filename
